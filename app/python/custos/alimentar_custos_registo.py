#!/usr/bin/env python3
"""
Une custos_linhas + facturas_extraidas + alocacao_diaria num ficheiro único
custos_registo.xlsx para controlo de custos por obra.
"""
import csv
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
DADOS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "dados"
CONFIG_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "config"
CUSTOS_LINHAS = DADOS_PATH / "custos_linhas.xlsx"
CUSTOS_REGISTO = DADOS_PATH / "custos_registo.xlsx"
ALOCACAO_DIARIA = DADOS_PATH / "alocacao_diaria.xlsx"
FACTURAS_EXTRAIDAS = DADOS_PATH / "facturas_extraidas"

COLUNAS = [
    "line_id",
    "document_no",
    "date",
    "supplier",
    "description",
    "quantity",
    "unit_price",
    "net_amount",
    "tax_pct",
    "tipo_linha",
    "centro_custo_codigo",
    "capitulo_orcamento",
    "origem",
]

TIPOS_CUSTO = ("subempreitadas", "materiais", "mao_obra", "equipamentos_maquinaria", "custos_sede")


def _load_csv_config(name: str, key_col: str, tipo_col: str) -> list[tuple[str, str]]:
    path = CONFIG_PATH / name
    if not path.exists():
        return []
    rows = []
    with open(path, encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            k = (row.get(key_col) or "").strip().lower()
            v = (row.get(tipo_col) or "").strip().lower()
            if k and v and v in TIPOS_CUSTO:
                rows.append((k, v))
    return rows


def _classificar_tipo(supplier: str, descricao: str, fornecedores: list, keywords: list) -> str:
    sup = (supplier or "").strip().lower()
    desc = (descricao or "").strip().lower()
    for key, tipo in fornecedores:
        if key in sup:
            return tipo
    for key, tipo in keywords:
        if key in desc:
            return tipo
    return "materiais"


def _load_sheet(path: Path) -> list[dict]:
    if not path.exists():
        return []
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    return [
        {h: ws.cell(r, i + 1).value for i, h in enumerate(headers) if h}
        for r in range(2, ws.max_row + 1)
    ]


def _linhas_de_custos(fornecedores: list, keywords: list) -> list[dict]:
    rows = []
    for r in _load_sheet(CUSTOS_LINHAS):
        cc = str(r.get("centro_custo_codigo") or "").strip()
        if not cc:
            continue
        tipo = (r.get("tipo_linha") or "materiais").strip().lower()
        if tipo not in TIPOS_CUSTO:
            tipo = "materiais"
        rows.append({
            "line_id": r.get("line_id"),
            "document_no": r.get("document_no"),
            "date": r.get("date"),
            "supplier": r.get("supplier"),
            "description": r.get("description"),
            "quantity": r.get("quantity"),
            "unit_price": r.get("unit_price"),
            "net_amount": r.get("net_amount"),
            "tax_pct": r.get("tax_pct"),
            "tipo_linha": tipo,
            "centro_custo_codigo": cc,
            "capitulo_orcamento": "",
            "origem": "compras",
        })
    return rows


def _linhas_de_facturas(fornecedores: list, keywords: list) -> list[dict]:
    rows = []
    if not FACTURAS_EXTRAIDAS.exists():
        return rows
    for jpath in FACTURAS_EXTRAIDAS.glob("*.json"):
        try:
            with open(jpath, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue
        centro = data.get("centro_custo_sugerido", "")
        if not centro and "|centro:" in (data.get("origem") or ""):
            centro = data["origem"].split("|centro:")[-1].strip()
        fornecedor = (data.get("fornecedor") or {}).get("nome", "")
        doc = data.get("documento") or {}
        doc_no = doc.get("numero", "")
        doc_date = doc.get("data", "")
        for i, ln in enumerate(data.get("linhas") or []):
            desc = ln.get("designacao", "")
            tipo = _classificar_tipo(fornecedor, desc, fornecedores, keywords)
            line_id = f"factura_{jpath.stem}_{i+1}"
            rows.append({
                "line_id": line_id,
                "document_no": doc_no,
                "date": doc_date,
                "supplier": fornecedor,
                "description": desc,
                "quantity": ln.get("quantidade"),
                "unit_price": ln.get("preco_unitario"),
                "net_amount": ln.get("valor_liquido"),
                "tax_pct": ln.get("iva_pct"),
                "tipo_linha": tipo,
                "centro_custo_codigo": centro,
                "capitulo_orcamento": "",
                "origem": "email",
            })
    return rows


def _linhas_de_alocacao() -> list[dict]:
    rows = []
    for r in _load_sheet(ALOCACAO_DIARIA):
        cc = str(r.get("centro_custo_codigo") or "").strip()
        if not cc:
            continue
        data_val = r.get("data")
        trab = r.get("trabalhador_codigo", "")
        horas = r.get("horas")
        notas = r.get("notas", "")
        line_id = f"aloc_{cc}_{data_val}_{trab}".replace(" ", "_") if data_val and trab else None
        if not line_id:
            continue
        rows.append({
            "line_id": line_id,
            "document_no": "alocacao",
            "date": data_val,
            "supplier": trab,
            "description": notas or "Mão de obra",
            "quantity": horas,
            "unit_price": None,
            "net_amount": None,
            "tax_pct": None,
            "tipo_linha": "mao_obra",
            "centro_custo_codigo": cc,
            "capitulo_orcamento": "",
            "origem": "alocacao",
        })
    return rows


def main() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    fornecedores = _load_csv_config("classificacao_fornecedores.csv", "fornecedor", "tipo")
    keywords = _load_csv_config("classificacao_keywords.csv", "keyword", "tipo")

    todas = []
    todas.extend(_linhas_de_custos(fornecedores, keywords))
    todas.extend(_linhas_de_facturas(fornecedores, keywords))
    todas.extend(_linhas_de_alocacao())

    DADOS_PATH.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "custos_registo"
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(COLUNAS, 1):
        c = ws.cell(1, col, h)
        c.fill = header_fill
        c.font = header_font
    for r, row in enumerate(todas, 2):
        for col, h in enumerate(COLUNAS, 1):
            ws.cell(r, col, value=row.get(h))

    wb.save(CUSTOS_REGISTO)
    print(f"✅ {CUSTOS_REGISTO} ({len(todas)} linhas)")
    por_origem = {}
    for row in todas:
        o = row.get("origem", "?")
        por_origem[o] = por_origem.get(o, 0) + 1
    for o, n in sorted(por_origem.items(), key=lambda x: -x[1]):
        print(f"   {o}: {n}")


if __name__ == "__main__":
    main()
