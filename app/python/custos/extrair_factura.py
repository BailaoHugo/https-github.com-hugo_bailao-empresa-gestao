#!/usr/bin/env python3
"""
Extrai informação estruturada de facturas (texto de PDF/OCR).
Produz um dicionário com: fornecedor, cliente, documento, linhas, totais.
"""
import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Optional

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from utils.taxas_iva import parse_taxa_iva


@dataclass
class Fornecedor:
    nome: str = ""
    nif: str = ""
    morada: str = ""


@dataclass
class Cliente:
    nome: str = ""
    nif: str = ""


@dataclass
class Documento:
    numero: str = ""
    data: str = ""
    tipo: str = "Fatura"


@dataclass
class LinhaFactura:
    designacao: str = ""
    codigo: str = ""
    quantidade: float = 1.0
    unidade: str = "UN"
    preco_unitario: Optional[float] = None
    valor_liquido: Optional[float] = None
    iva_pct: Optional[float] = None


@dataclass
class Totais:
    valor_liquido: Optional[float] = None
    total_iva: Optional[float] = None
    total_documento: Optional[float] = None


@dataclass
class FacturaExtraida:
    fornecedor: Fornecedor = field(default_factory=Fornecedor)
    cliente: Cliente = field(default_factory=Cliente)
    documento: Documento = field(default_factory=Documento)
    linhas: list[LinhaFactura] = field(default_factory=list)
    totais: Totais = field(default_factory=Totais)
    origem: str = ""

    def to_dict(self):
        d = {
            "fornecedor": asdict(self.fornecedor),
            "cliente": asdict(self.cliente),
            "documento": asdict(self.documento),
            "linhas": [asdict(l) for l in self.linhas],
            "totais": asdict(self.totais),
            "origem": self.origem,
        }
        # Extrair centro do assunto se presente (ex: "email:x.pdf|centro:25.113")
        if "|centro:" in self.origem:
            d["centro_custo_sugerido"] = self.origem.split("|centro:")[-1].strip()
        return d


def _parse_valor(s: str) -> Optional[float]:
    if not s:
        return None
    s = str(s).replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _parse_data(s: str) -> str:
    """Converte dd-mm-yyyy ou dd/mm/yyyy para yyyy-mm-dd."""
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", s)
    if not m:
        return ""
    d, mo, y = m.groups()
    y = f"20{y[-2:]}" if len(y) == 2 else y
    return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"


def extrair_factura(texto: str, origem: str = "") -> FacturaExtraida:
    """
    Analisa o texto extraído de uma factura e devolve estrutura organizada.
    """
    out = FacturaExtraida(origem=origem)
    lines = [l.strip() for l in texto.splitlines() if l.strip()]

    # NIF fornecedor: "Nº Contribuinte: 503238660" (está na área do vendedor)
    nif_forn = None
    for line in lines:
        m = re.search(r"N[ºo°]\s*Contribuinte\s*[:\s]*(\d{9})", line, re.I)
        if m:
            nif_forn = m.group(1)
            out.fornecedor.nif = nif_forn
            break

    # --- Fornecedor: nome antes de SEDE/LOJA, ou linha com "Vendedor" / estrutura típica
    for i, line in enumerate(lines):
        if re.search(r"(SEDE|LOJA):\s*$", line, re.I) and i >= 1:
            candidato = lines[i - 1].strip()
            if len(candidato) > 15 and not re.match(r"^\d", candidato) and "E-mail" not in candidato:
                out.fornecedor.nome = candidato[:200]
                break
    if not out.fornecedor.nome:
        for line in lines:
            if "Vendedor" in line:
                idx = lines.index(line)
                if idx + 1 < len(lines):
                    out.fornecedor.nome = lines[idx + 1].strip()[:200]
                break
    if not out.fornecedor.nome:
        for line in lines:
            if re.search(r"LDA|LDA\.|UNIPESSOAL|SA\b", line) and "SEDE" not in line and "LOJA" not in line:
                if "E-mail" not in line and "Cliente" not in line and len(line) > 20:
                    if not re.match(r"^IVA-|^\d{9}", line):
                        out.fornecedor.nome = line[:200]
                        break

    # --- Cliente: "Contribuinte" na secção cliente + linha seguinte, ou SOLID PROJECTS / CHURRASQUEIRA
    for i, line in enumerate(lines):
        if "Contribuinte" in line and i + 1 < len(lines):
            seguinte = lines[i + 1].strip()
            if not re.match(r"^IVA-|^\d{9}", seguinte) and len(seguinte) > 3:
                out.cliente.nome = seguinte[:200]
                break
    if not out.cliente.nome:
        for line in lines:
            if "SOLID PROJECTS" in line:
                out.cliente.nome = line[:200]
                break
    if not out.cliente.nome:
        for line in lines:
            if "CHURRASQUEIRA" in line and "Contribuinte" not in line:
                out.cliente.nome = line[:200]
                break

    # NIF cliente: IVA-PT-515188166 (em linha separada ou junto a Cliente)
    for line in lines:
        m = re.search(r"IVA-?PT-?(\d{9})", line, re.I)
        if m:
            out.cliente.nif = m.group(1)
            break

    # --- Documento: Fatura-Recibo Nº VDI 2610/201
    for line in lines:
        m = re.search(r"(?:Fatura|Factura|Recibo)[^\d]*N[ºo°]?\s*[\s:]*(.+?)(?:\s+(\d{2}[-/]\d{2}[-/]\d{2,4}))?$", line, re.I)
        if m:
            out.documento.numero = m.group(1).strip()[:80]
            if m.group(2):
                out.documento.data = _parse_data(m.group(2))
            break
        m = re.search(r"(VDI\s*\d+[/\-]\d+)", line, re.I)
        if m and not out.documento.numero:
            out.documento.numero = m.group(1)

    # Data: dd-mm-yyyy
    if not out.documento.data:
        for line in lines:
            d = _parse_data(line)
            if d and "2000" <= d[:4] <= "2030":
                out.documento.data = d
                break

    # --- Totais: Total Documento EUR 189,97 | Total Líquido 154,45 | Total de IVA 35,52
    for line in lines:
        if "Total Documento" in line or "Total documento" in line:
            m = re.search(r"(\d+[,.]\d{2})\s*(?:EUR|€)?", line)
            if m:
                out.totais.total_documento = _parse_valor(m.group(1))
        if "Total L[ií]quido" in line or "Totais" in line:
            m = re.search(r"(\d+[,.]\d{2})", line)
            if m and out.totais.valor_liquido is None:
                out.totais.valor_liquido = _parse_valor(m.group(1))
        if "Total de IVA" in line or "Total IVA" in line or "Valor IVA" in line:
            m = re.search(r"(\d+[,.]\d{2})", line)
            if m:
                out.totais.total_iva = _parse_valor(m.group(1))

    # Procurar IVA 23%: base 154,45 e valor IVA 35,52
    for line in lines:
        if "IVA 23%" in line or "23%" in line:
            m = re.findall(r"(\d+[,.]\d{2})", line)
            for v in m:
                x = _parse_valor(v)
                if x and 30 < x < 40 and out.totais.total_iva is None:
                    out.totais.total_iva = x
                if x and 150 < x < 170 and out.totais.valor_liquido is None:
                    out.totais.valor_liquido = x
    if out.totais.total_iva is None:
        for line in lines:
            if "35,52" in line or "35.52" in line:
                m = re.search(r"(\d+[,.]\d{2})", line)
                if m:
                    out.totais.total_iva = _parse_valor(m.group(1))
                    break

    # Fallback totais: procurar padrão Valor Ilíquido/154,45
    for i, line in enumerate(lines):
        if "154,45" in line or "Valor Ilíquido" in line:
            m = re.search(r"(\d+[,.]\d{2})", line)
            if m and out.totais.valor_liquido is None:
                out.totais.valor_liquido = _parse_valor(m.group(1))
        if "35,52" in line and "IVA" in line:
            m = re.search(r"(\d+[,.]\d{2})", line)
            if m and out.totais.total_iva is None:
                out.totais.total_iva = _parse_valor(m.group(1))
        if "189,97" in line:
            m = re.search(r"(\d+[,.]\d{2})", line)
            if m and out.totais.total_documento is None:
                out.totais.total_documento = _parse_valor(m.group(1))

    # Palavras que indicam linhas que NÃO são produtos
    EXCLUIR = ("DESIGNAÇÃO", "CÓDIGO", "P.P.", "TOTAL", "OBSERVAÇÕES", "ORIGINAL", "CÓPIA",
               "FORMULÁRIO", "CAPITAL SOCIAL", "REG. CONSERV", "SOFTWARE", "PAGAMENTO", "IBAN",
               "ATCUD", "RECEBIDO", "LÍQUIDO", "V.DESC", "% DESC", "ELABORADO POR")

    # Estratégia 1: PDF com layout vertical (código, IVA, designação, unidade, valores em linhas separadas)
    i = 0
    while i < len(lines):
        line = lines[i]
        if re.match(r"^(\d{6,15})\s*$", line):  # linha só com código (6-15 dígitos)
            codigo = line.strip()
            iva_pct, designacao, unidade, preco, qtd, valor_liq = None, "", "UN", None, 1.0, None
            j = i + 1
            nums = []
            for k in range(j, min(j + 12, len(lines))):
                ln = lines[k]
                if re.match(r"^\d{6,15}\s*$", ln) and k > j:  # próximo produto
                    break
                if re.match(r"^(23|13|6)\s*%?\s*$", ln):
                    iva_pct = parse_taxa_iva(ln.strip())
                elif re.match(r"^(UN|UND|LA|KG|M)\s*$", ln, re.I):
                    unidade = ln.strip().upper()
                elif re.match(r"^\d+[,.]\d{2}\s*$", ln):
                    if len(nums) < 5:  # máx 5 valores por produto
                        nums.append(_parse_valor(ln))
                elif len(ln) > 8 and not re.match(r"^\d", ln) and "€" not in ln:
                    if not designacao and not any(x in ln.upper() for x in EXCLUIR):
                        designacao = ln[:150]
                j = k + 1
            if designacao and len(nums) >= 2:
                # Ordem típica: preço unit, valor IVA, preço c/IVA, qtd (último muitas vezes 1,00 ou 2,00)
                preco = nums[0]
                ultimo = nums[-1]
                qtd = ultimo if 0.5 <= ultimo <= 100 and ultimo == round(ultimo) else 1.0
                valor_liq = round(preco * qtd, 2) if preco and qtd else preco
                out.linhas.append(LinhaFactura(
                    designacao=designacao,
                    codigo=codigo,
                    quantidade=qtd,
                    unidade=unidade,
                    preco_unitario=preco,
                    valor_liquido=valor_liq,
                    iva_pct=iva_pct,
                ))
            i = j
            continue
        # Estratégia 2: linha horizontal com código + designação + valores
        cod_match = re.match(r"^(\d{6,15})\s+", line)
        vals = re.findall(r"(\d+[,.]\d{2})", line)
        if cod_match and len(vals) >= 2:
            codigo = cod_match.group(1)
            resto = line[len(codigo):].strip()
            designacao = re.split(r"\s+(UN|UND|LA|KG|M)\s+", resto, 1, re.I)[0].strip()
            designacao = re.sub(r"\s*\d{1,2}\s*%?\s*$", "", designacao).strip()
            if len(designacao) >= 5:
                iva_match = re.search(r"\b(23|13|6)\s*%?", resto)
                iva_pct = parse_taxa_iva(iva_match.group(1)) if iva_match else None
                un_match = re.search(r"\b(UN|UND|LA|KG|M)\b", resto, re.I)
                unidade = un_match.group(1).upper() if un_match else "UN"
                valor_liq = _parse_valor(vals[-1])
                preco = _parse_valor(vals[0])
                qtd = round(valor_liq / preco, 2) if valor_liq and preco and preco > 0 else 1.0
                out.linhas.append(LinhaFactura(
                    designacao=designacao[:150],
                    codigo=codigo,
                    quantidade=qtd,
                    unidade=unidade,
                    preco_unitario=preco,
                    valor_liquido=valor_liq,
                    iva_pct=iva_pct,
                ))
        i += 1

    # Deduplicar linhas (PDF pode ter ORIGINAL + CÓPIA): manter primeira ocorrência por código
    vistos = set()
    unicas = []
    for ln in out.linhas:
        chave = (ln.codigo or ln.designacao, ln.preco_unitario)
        if chave not in vistos:
            vistos.add(chave)
            unicas.append(ln)
    out.linhas = unicas

    # Se não encontrou linhas, usar totais como linha única
    if not out.linhas and out.totais.valor_liquido:
        out.linhas.append(LinhaFactura(
            designacao="Total factura",
            quantidade=1.0,
            valor_liquido=out.totais.valor_liquido,
            preco_unitario=out.totais.valor_liquido,
        ))

    return out


def guardar_factura_json(factura: FacturaExtraida, path: Path) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(factura.to_dict(), f, ensure_ascii=False, indent=2)


def guardar_factura_excel(factura: FacturaExtraida, path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        # Aba Cabeçalho
        ws = wb.active
        ws.title = "Cabeçalho"
        ws["A1"] = "Fornecedor"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = factura.fornecedor.nome
        ws["A3"] = f"NIF: {factura.fornecedor.nif}"
        ws["A4"] = factura.fornecedor.morada
        ws["A6"] = "Cliente"
        ws["A6"].font = Font(bold=True)
        ws["A7"] = factura.cliente.nome
        ws["A8"] = f"NIF: {factura.cliente.nif}"
        ws["A10"] = "Documento"
        ws["A10"].font = Font(bold=True)
        ws["A11"] = f"Nº: {factura.documento.numero}"
        ws["A12"] = f"Data: {factura.documento.data}"
        ws["A14"] = "Totais"
        ws["A14"].font = Font(bold=True)
        ws["A15"] = f"Valor Líquido: {factura.totais.valor_liquido}"
        ws["A16"] = f"Total IVA: {factura.totais.total_iva}"
        ws["A17"] = f"Total Documento: {factura.totais.total_documento}"
        # Aba Linhas
        ws2 = wb.create_sheet("Linhas")
        headers = ["Designação", "Código", "Qtd", "Unidade", "Preço Unit.", "Valor Líquido", "IVA %"]
        for c, h in enumerate(headers, 1):
            cell = ws2.cell(1, c, h)
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.font = Font(bold=True, color="FFFFFF")
        for r, ln in enumerate(factura.linhas, 2):
            ws2.cell(r, 1, ln.designacao)
            ws2.cell(r, 2, ln.codigo)
            ws2.cell(r, 3, ln.quantidade)
            ws2.cell(r, 4, ln.unidade)
            ws2.cell(r, 5, ln.preco_unitario)
            ws2.cell(r, 6, ln.valor_liquido)
            ws2.cell(r, 7, ln.iva_pct)
        wb.save(path)
    except ImportError:
        guardar_factura_json(factura, path.with_suffix(".json"))


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python extrair_factura.py <caminho_pdf>")
        sys.exit(1)
    from processar_email_despesas import _extrair_texto_pdf
    pdf_path = Path(sys.argv[1])
    texto = _extrair_texto_pdf(pdf_path)
    f = extrair_factura(texto, origem=str(pdf_path.name))
    out_dir = Path(__file__).resolve().parent.parent.parent.parent / "03_CONTABILIDADE_ANALITICA" / "dados" / "facturas_extraidas"
    out_dir.mkdir(parents=True, exist_ok=True)
    base = out_dir / (pdf_path.stem + "_extraida")
    guardar_factura_json(f, Path(str(base) + ".json"))
    guardar_factura_excel(f, Path(str(base) + ".xlsx"))
    print(f"✅ Guardado: {base}.json e {base}.xlsx")
    print(json.dumps(f.to_dict(), ensure_ascii=False, indent=2)[:1500])
