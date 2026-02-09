#!/usr/bin/env python3
"""
Processa emails enviados para registardespesa@ennova.pt:
- Assunto: centro de custo (ex: "001" ou "054 - Estoril 124")
- Anexo: foto da fatura/recibo
- Faz OCR, extrai dados, grava em custos_linhas.xlsx
- Opcional: executa exportar_custos_por_obra para gerar Excel por obra

Variáveis de ambiente:
  EMAIL_IMAP_HOST=ennova.pt
  EMAIL_IMAP_PORT=993
  EMAIL_USER=registardespesa@ennova.pt
  EMAIL_PASSWORD=...
  GESTAO_BASE_PATH=/path/to/GESTAO_EMPRESA
"""
import email
import imaplib
import os
import re
import sys
from email.header import decode_header
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from utils.taxas_iva import parse_taxa_iva

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
DADOS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "dados"
CONFIG_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "config"
CUSTOS_LINHAS = DADOS_PATH / "custos_linhas.xlsx"
CUSTOS_REGISTO = DADOS_PATH / "custos_registo.xlsx"
UPLOADS_PATH = DADOS_PATH / "uploads"
UPLOADS_PATH.mkdir(parents=True, exist_ok=True)

try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    XL_AVAILABLE = True
except ImportError:
    XL_AVAILABLE = False

# Extensões aceites: imagens e PDF
IMAGE_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tiff"}
PDF_EXT = {".pdf"}
ANEXO_EXT = IMAGE_EXT | PDF_EXT

try:
    import fitz  # PyMuPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


def _decode_header_value(header_val):
    if header_val is None:
        return ""
    if isinstance(header_val, str):
        return header_val.strip()
    parts = []
    for part, charset in decode_header(header_val):
        if isinstance(part, bytes):
            parts.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            parts.append(str(part))
    return " ".join(parts).strip()


def _extrair_centro_assunto(assunto: str, centros_validos: set[str]) -> str | None:
    """
    Extrai o código do centro de custo do assunto.
    Formatos aceites:
      - YY.NNN: "25.113", "25.113 - CCG" (2 dig=ano, 3 dig=obra)
      - Antigo: "001", "054", "[001]", "001 - Sede"
    """
    assunto = assunto.strip()
    if not assunto:
        return None
    aceitar_qualquer = len(centros_validos) == 0

    def _validar(cod: str) -> bool:
        if cod in centros_validos:
            return True
        if aceitar_qualquer:
            # YY.NN ou YY.NNN
            if re.match(r"^\d{2}\.\d{2,3}$", cod):
                return True
            if len(cod) == 3 and cod.isdigit():
                return True
        return False

    # Formato YY.NNN (ex: 25.113, 24.54, 99.990)
    m = re.search(r"(\d{2}\.\d{2,3})\b", assunto)
    if m:
        cod = m.group(1)
        if _validar(cod):
            return cod
    # Tentar padrão [CODIGO]
    m = re.search(r"\[(\d{3}|\d{2}\.\d{2,3})\]", assunto)
    if m:
        cod = m.group(1)
        if _validar(cod):
            return cod
    # Primeiro token YY.NNN ou 3 dígitos
    tokens = re.split(r"[\s\-–—]+", assunto)
    for t in tokens:
        t = t.strip()
        if re.match(r"^\d{2}\.\d{2,3}$", t) and _validar(t):
            return t
        if len(t) >= 3 and t[:3].isdigit() and _validar(t[:3]):
            return t[:3]
        if _validar(t):
            return t
    return None


def _carregar_centros_validos() -> set[str]:
    path = BASE_PATH / "00_CONFIG" / "centros_custo.xlsx"
    if not path.exists() or not XL_AVAILABLE:
        return set()
    try:
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            idx = next((i for i, c in enumerate(ws[1], 1) if c.value == "centro_custo_codigo"), None)
            if idx:
                return {str(ws.cell(r, idx).value).strip() for r in range(2, ws.max_row + 1) if ws.cell(r, idx).value}
        return set()
    except Exception:
        return set()


def _ocr_image(img_path: Path) -> str:
    if not OCR_AVAILABLE:
        return ""
    try:
        img = Image.open(img_path)
        if img.mode != "RGB":
            img = img.convert("RGB")
        return pytesseract.image_to_string(img, lang="por+eng")
    except Exception as e:
        return f"[OCR erro: {e}]"


def _extrair_texto_pdf(pdf_path: Path) -> str:
    """
    Extrai texto do PDF: primeiro tenta texto embutido (PDF digital).
    Se pouco texto, converte páginas a imagem e faz OCR (PDF digitalizado).
    """
    if not PDF_AVAILABLE:
        return ""
    try:
        doc = fitz.open(pdf_path)
        texto = ""
        for i in range(len(doc)):
            texto += doc[i].get_text()
        doc.close()
        # Se tem texto suficiente, usar direto
        if len(texto.strip()) >= 50:
            return texto
        # PDF digitalizado: converter páginas a imagem e OCR
        if not OCR_AVAILABLE:
            return texto or "[PDF sem texto embutido; instale pytesseract para OCR]"
        doc = fitz.open(pdf_path)
        partes = []
        for i in range(len(doc)):
            page = doc[i]
            pix = page.get_pixmap(dpi=150)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            partes.append(pytesseract.image_to_string(img, lang="por+eng"))
        doc.close()
        return "\n".join(partes)
    except Exception as e:
        return f"[PDF erro: {e}]"


def _extrair_dados_ocr(texto: str) -> dict:
    out = {"supplier": "", "date": "", "net_amount": None, "tax_pct": None, "description": ""}
    lines = [l.strip() for l in texto.splitlines() if l.strip()]
    for line in lines:
        m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", line)
        if m and not out["date"]:
            d, mo, y = m.groups()
            out["date"] = f"20{y[-2:]}-{mo.zfill(2)}-{d.zfill(2)}" if len(y) == 2 else f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
        for p in [r"total[:\s]*(\d+[,.]\d{2})", r"(\d+[,.]\d{2})\s*€", r"iva[:\s]*(\d+[,.]?\d*)%?"]:
            m = re.search(p, line.lower())
            if m:
                v = m.group(1).replace(",", ".")
                try:
                    n = float(v)
                    if "iva" in line.lower() or "%" in line:
                        out["tax_pct"] = parse_taxa_iva(v)
                    elif out["net_amount"] is None and n > 0:
                        out["net_amount"] = n
                except ValueError:
                    pass
    if lines and not out["supplier"]:
        out["supplier"] = lines[0][:100]
    return out


def _append_custo(centro: str, dados: dict, origem: str = "email") -> None:
    if not XL_AVAILABLE:
        return
    if not CUSTOS_LINHAS.exists():
        DADOS_PATH.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "custos_linhas"
        headers = ["line_id", "document_id", "document_no", "date", "supplier", "description",
                   "quantity", "unit_price", "net_amount", "tax_pct", "tipo_linha", "centro_custo_codigo"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.font = Font(bold=True, color="FFFFFF")
        wb.save(CUSTOS_LINHAS)
    wb = load_workbook(CUSTOS_LINHAS)
    ws = wb.active
    next_row = ws.max_row + 1
    row_data = {
        "line_id": f"email_{next_row}",
        "document_id": "",
        "document_no": origem,
        "date": dados.get("date"),
        "supplier": dados.get("supplier"),
        "description": dados.get("description") or f"Registo por {origem}",
        "quantity": 1,
        "unit_price": dados.get("net_amount"),
        "net_amount": dados.get("net_amount"),
        "tax_pct": dados.get("tax_pct"),
        "tipo_linha": "materiais",
        "centro_custo_codigo": centro,
    }
    headers = [c.value for c in ws[1]]
    for col, h in enumerate(headers, 1):
        if h in row_data:
            ws.cell(row=next_row, column=col, value=row_data[h])
    wb.save(CUSTOS_LINHAS)


CUSTOS_REGISTO_COLUNAS = [
    "line_id", "document_no", "date", "supplier", "description", "quantity",
    "unit_price", "net_amount", "tax_pct", "tipo_linha", "centro_custo_codigo",
    "capitulo_orcamento", "origem",
]


def _append_factura_to_custos_registo(factura, centro: str, base_name: str) -> None:
    """Insere linhas da factura extraída em custos_registo.xlsx."""
    if not XL_AVAILABLE:
        return
    import csv
    fornecedores, keywords = [], []
    for path, kc, tc in [
        (CONFIG_PATH / "classificacao_fornecedores.csv", "fornecedor", "tipo"),
        (CONFIG_PATH / "classificacao_keywords.csv", "keyword", "tipo"),
    ]:
        if path.exists():
            with open(path, encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    k = (row.get(kc) or "").strip().lower()
                    v = (row.get(tc) or "").strip().lower()
                    if k and v:
                        (fornecedores if kc == "fornecedor" else keywords).append((k, v))
    sup = (factura.fornecedor.nome or "").lower()
    def _tipo(desc):
        desc = (desc or "").lower()
        for k, t in fornecedores:
            if k in sup:
                return t
        for k, t in keywords:
            if k in desc:
                return t
        return "materiais"

    DADOS_PATH.mkdir(parents=True, exist_ok=True)
    if not CUSTOS_REGISTO.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "custos_registo"
        for c, h in enumerate(CUSTOS_REGISTO_COLUNAS, 1):
            cell = ws.cell(1, c, h)
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.font = Font(bold=True, color="FFFFFF")
        wb.save(CUSTOS_REGISTO)
    wb = load_workbook(CUSTOS_REGISTO)
    ws = wb.active
    doc_no = factura.documento.numero
    doc_date = factura.documento.data
    fornecedor = factura.fornecedor.nome
    for i, ln in enumerate(factura.linhas):
        row = ws.max_row + 1
        tipo = _tipo(ln.designacao)
        line_id = f"email_{base_name}_{i+1}"
        vals = [
            line_id, doc_no, doc_date, fornecedor, ln.designacao,
            ln.quantidade, ln.preco_unitario, ln.valor_liquido, ln.iva_pct,
            tipo, centro, "", "email",
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row, col, value=v)
    wb.save(CUSTOS_REGISTO)


def processar_email() -> int:
    """
    Conecta ao IMAP, processa emails não lidos, extrai anexos de imagem,
    faz OCR e grava em custos_linhas. Marca emails como lidos.
    Retorna o número de despesas processadas.
    """
    host = os.getenv("EMAIL_IMAP_HOST", "ennova.pt")
    port = int(os.getenv("EMAIL_IMAP_PORT", "993"))
    user = os.getenv("EMAIL_USER", "registardespesa@ennova.pt")
    password = os.getenv("EMAIL_PASSWORD", "")
    if not password:
        print("AVISO: EMAIL_PASSWORD não definido. Use variável de ambiente.")
        return 0

    centros = _carregar_centros_validos()
    processados = 0

    try:
        mail = imaplib.IMAP4_SSL(host, port)
        mail.login(user, password)
        mail.select("INBOX")
        typ, data = mail.search(None, "UNSEEN")
        if typ != "OK" or not data[0]:
            mail.logout()
            return 0

        for num in data[0].split():
            typ, msg_data = mail.fetch(num, "(RFC822)")
            if typ != "OK":
                continue
            raw = msg_data[0][1]
            msg = email.message_from_bytes(raw)
            assunto = _decode_header_value(msg.get("Subject", ""))

            centro = _extrair_centro_assunto(assunto, centros)
            if not centro:
                print(f"  Ignorado (assunto sem centro válido): {assunto[:60]}")
                continue

            # Procurar anexos (imagem ou PDF)
            for part in msg.walk():
                if part.get_content_maintype() == "multipart":
                    continue
                fname = part.get_filename()
                if not fname:
                    continue
                fname = _decode_header_value(fname)
                ext = Path(fname).suffix.lower()
                if ext not in ANEXO_EXT:
                    continue

                payload = part.get_payload(decode=True)
                if not payload:
                    continue

                save_path = UPLOADS_PATH / f"email_{os.urandom(6).hex()}{ext}"
                save_path.write_bytes(payload)

                if ext in IMAGE_EXT:
                    texto = _ocr_image(save_path)
                elif ext in PDF_EXT:
                    texto = _extrair_texto_pdf(save_path)
                else:
                    texto = ""

                # Extrair factura completa para ficheiro organizado
                from custos.extrair_factura import (
                    extrair_factura,
                    guardar_factura_json,
                    guardar_factura_excel,
                )
                factura = extrair_factura(texto, origem=f"email:{fname}|centro:{centro}")
                factura_extras_dir = DADOS_PATH / "facturas_extraidas"
                factura_extras_dir.mkdir(parents=True, exist_ok=True)
                base_name = save_path.stem
                guardar_factura_json(factura, factura_extras_dir / f"{base_name}.json")
                guardar_factura_excel(factura, factura_extras_dir / f"{base_name}.xlsx")

                # Inserir linhas em custos_registo
                _append_factura_to_custos_registo(factura, centro, base_name)

                # Opcional: append a custos_linhas (legado)
                if os.getenv("APPEND_CUSTOS", "0") == "1":
                    dados = _extrair_dados_ocr(texto)
                    dados["description"] = factura.documento.numero or fname
                    dados["supplier"] = factura.fornecedor.nome or dados.get("supplier")
                    dados["date"] = factura.documento.data or dados.get("date")
                    dados["net_amount"] = factura.totais.valor_liquido or factura.totais.total_documento
                    dados["tax_pct"] = factura.linhas[0].iva_pct if factura.linhas else dados.get("tax_pct")
                    _append_custo(centro, dados, origem=f"email:{fname}")

                processados += 1
                print(f"  ✅ Processado: centro={centro} | {fname} -> facturas_extraidas + custos_registo")

            mail.store(num, "+FLAGS", "\\Seen")

        mail.logout()
    except imaplib.IMAP4.error as e:
        print(f"Erro IMAP: {e}")
    except Exception as e:
        print(f"Erro: {e}")
        raise

    return processados


def main():
    print("Processar emails de despesas (registardespesa@ennova.pt)...")
    n = processar_email()
    print(f"\n✅ {n} despesa(s) processada(s) -> facturas_extraidas + custos_registo.xlsx")

    if n > 0 and os.getenv("EXPORTAR_POR_OBRA", "0") == "1":
        print("\nA executar exportar_custos_por_obra...")
        try:
            from custos.exportar_custos_por_obra import main as export_main
            export_main()
        except Exception as e:
            print(f"  Aviso: {e}")


if __name__ == "__main__":
    main()
