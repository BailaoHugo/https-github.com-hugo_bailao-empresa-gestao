#!/usr/bin/env python3
"""
Cria a estrutura de trabalhadores e alocação diária.
- trabalhadores.xlsx: codigo, nome, origem (toconline|externo)
- alocacao_diaria.xlsx: data, trabalhador_codigo, centro_custo_codigo, horas, notas
"""
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
DADOS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "dados"
CENTROS = BASE_PATH / "00_CONFIG" / "centros_custo.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _load_centros_codigos() -> list[str]:
    if not CENTROS.exists():
        return []
    wb = load_workbook(CENTROS, data_only=True)
    ws = wb.active
    idx = next((i for i, c in enumerate(ws[1], 1) if c.value == "centro_custo_codigo"), None)
    if not idx:
        return []
    return [str(ws.cell(r, idx).value or "").strip() for r in range(2, ws.max_row + 1) if ws.cell(r, idx).value]


def main() -> None:
    DADOS_PATH.mkdir(parents=True, exist_ok=True)
    centros = _load_centros_codigos()

    # trabalhadores.xlsx
    path_trab = DADOS_PATH / "trabalhadores.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "trabalhadores"
    headers = ["codigo", "nome", "origem"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    ws.cell(row=2, column=1, value="EXT001")
    ws.cell(row=2, column=2, value="Exemplo Externo")
    ws.cell(row=2, column=3, value="externo")
    wb.save(path_trab)
    print(f"✅ {path_trab} (adicione trabalhadores do TOCOnline e externos)")

    # alocacao_diaria.xlsx
    path_aloc = DADOS_PATH / "alocacao_diaria.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "alocacao_diaria"
    headers2 = ["data", "trabalhador_codigo", "centro_custo_codigo", "horas", "notas"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    wb2.save(path_aloc)
    print(f"✅ {path_aloc} (registo diário: trabalhador + centro de custo)")
    print(f"   Centros disponíveis: {', '.join(centros[:5])}{'...' if len(centros) > 5 else ''}")


if __name__ == "__main__":
    main()
