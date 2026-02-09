#!/usr/bin/env python3
"""
Gera 00_CONFIG/capitulos_orcamento.xlsx a partir de catalogo.json.
Usado para validação/dropdown de capitulo_orcamento em custos_registo.
"""
import json
import os
from pathlib import Path

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
CATALOGO = BASE_PATH / "gestao-web" / "public" / "catalogo.json"
OUT_FILE = BASE_PATH / "00_CONFIG" / "capitulos_orcamento.xlsx"


def main() -> None:
    if not CATALOGO.exists():
        print(f"AVISO: {CATALOGO} não encontrado.")
        return
    with open(CATALOGO, encoding="utf-8") as f:
        data = json.load(f)
    capitulos = data.get("capitulos", [])
    if not capitulos:
        print("Nenhum capítulo no catálogo.")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "capitulos"
    headers = ["capitulo_id", "capitulo_nome", "area"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.font = Font(bold=True, color="FFFFFF")
    for r, cap in enumerate(sorted(capitulos, key=lambda x: (x.get("area", ""), x.get("ordem", 0))), 2):
        ws.cell(r, 1, cap.get("id", ""))
        ws.cell(r, 2, cap.get("nome", ""))
        ws.cell(r, 3, cap.get("area", ""))
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)
    print(f"✅ {OUT_FILE} ({len(capitulos)} capítulos)")


if __name__ == "__main__":
    main()
