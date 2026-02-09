#!/usr/bin/env python3
"""
Lê custos_registo.xlsx e gera um Excel por centro de custo (obra)
com 5 abas: subempreitadas, materiais, mao_obra, equipamentos_maquinaria, custos_sede.
Fonte: custos_registo (ou custos_linhas+alocacao se custos_registo não existir).
"""
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
DADOS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "dados"
OBRAS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "obras"
CUSTOS_REGISTO = DADOS_PATH / "custos_registo.xlsx"
CUSTOS_LINHAS = DADOS_PATH / "custos_linhas.xlsx"
ALOCACAO_DIARIA = DADOS_PATH / "alocacao_diaria.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Colunas com capitulo_orcamento
COLUNAS_COM_CAPITULO = [
    "date", "document_no", "supplier", "description", "capitulo_orcamento",
    "quantity", "unit_price", "net_amount", "tax_pct",
]
COLUNAS_MO = ["date", "supplier", "description", "capitulo_orcamento", "quantity", "notas"]

TIPOS = ("subempreitadas", "materiais", "mao_obra", "equipamentos_maquinaria", "custos_sede")


def _load_sheet(path: Path) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    return [
        {h: ws.cell(r, i + 1).value for i, h in enumerate(headers) if h}
        for r in range(2, ws.max_row + 1)
    ]


def _load_custos_registo() -> list[dict]:
    """Carrega custos_registo ou fallback para custos_linhas+alocacao."""
    if CUSTOS_REGISTO.exists():
        rows = _load_sheet(CUSTOS_REGISTO)
        return [r for r in rows if (r.get("centro_custo_codigo") or "").strip()]
    # Fallback: custos_linhas + alocacao
    result = []
    for r in _load_sheet(CUSTOS_LINHAS):
        cc = str(r.get("centro_custo_codigo") or "").strip()
        if not cc:
            continue
        row = dict(r)
        row["capitulo_orcamento"] = ""
        result.append(row)
    for r in _load_sheet(ALOCACAO_DIARIA):
        cc = str(r.get("centro_custo_codigo") or "").strip()
        if not cc:
            continue
        result.append({
            "line_id": f"aloc_{r.get('data')}_{r.get('trabalhador_codigo')}",
            "document_no": "alocacao",
            "date": r.get("data"),
            "supplier": r.get("trabalhador_codigo"),
            "description": r.get("notas") or "Mão de obra",
            "quantity": r.get("horas"),
            "unit_price": None,
            "net_amount": None,
            "tax_pct": None,
            "tipo_linha": "mao_obra",
            "centro_custo_codigo": cc,
            "capitulo_orcamento": "",
            "origem": "alocacao",
        })
    return result


def _mapear_para_aba(row: dict, tipo: str) -> dict:
    """Mapeia linha para formato da aba (com capitulo_orcamento)."""
    if tipo == "mao_obra":
        return {
            "date": row.get("date"),
            "supplier": row.get("supplier"),
            "description": row.get("description"),
            "capitulo_orcamento": row.get("capitulo_orcamento") or "",
            "quantity": row.get("quantity"),
            "notas": row.get("description"),
        }
    return {
        "date": row.get("date"),
        "document_no": row.get("document_no"),
        "supplier": row.get("supplier"),
        "description": row.get("description"),
        "capitulo_orcamento": row.get("capitulo_orcamento") or "",
        "quantity": row.get("quantity"),
        "unit_price": row.get("unit_price"),
        "net_amount": row.get("net_amount"),
        "tax_pct": row.get("tax_pct"),
    }


def _escrever_aba(ws, colunas: list[str], rows: list[dict]) -> None:
    for col, h in enumerate(colunas, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for r, row in enumerate(rows, 2):
        for col, h in enumerate(colunas, 1):
            ws.cell(row=r, column=col, value=row.get(h))


def main() -> None:
    rows = _load_custos_registo()
    if not rows:
        print("⚠️ Nenhuma linha com centro_custo_codigo.")
        print("   Execute alimentar_custos_registo.py primeiro.")
        return

    por_centro: dict[str, list[dict]] = {}
    for r in rows:
        cc = str(r.get("centro_custo_codigo", "")).strip()
        if cc not in por_centro:
            por_centro[cc] = []
        por_centro[cc].append(r)

    OBRAS_PATH.mkdir(parents=True, exist_ok=True)
    for cc, itens in por_centro.items():
        wb = Workbook()
        # Normalizar tipo_linha (subempreitada -> subempreitadas)
        tipo_map = {"subempreitada": "subempreitadas", "subempreitadas": "subempreitadas",
                    "materiais": "materiais", "mao_obra": "mao_obra",
                    "equipamentos_maquinaria": "equipamentos_maquinaria",
                    "custos_sede": "custos_sede"}
        por_tipo: dict[str, list] = {t: [] for t in TIPOS}
        for r in itens:
            t = (r.get("tipo_linha") or "materiais").strip().lower()
            t = tipo_map.get(t, "materiais" if t not in TIPOS else t)
            if t in por_tipo:
                mapped = _mapear_para_aba(r, t)
                por_tipo[t].append(mapped)

        first = True
        for tipo in TIPOS:
            linhas = por_tipo[tipo]
            if first:
                ws = wb.active
                first = False
            else:
                ws = wb.create_sheet(tipo)
            ws.title = tipo[:31]
            cols = COLUNAS_MO if tipo == "mao_obra" else COLUNAS_COM_CAPITULO
            _escrever_aba(ws, cols, linhas)

        path = OBRAS_PATH / f"custo_obra_{cc.replace('.', '_')}.xlsx"
        wb.save(path)
        counts = " | ".join(f"{t}: {len(por_tipo[t])}" for t in TIPOS)
        print(f"✅ {path} ({counts})")
    print(f"\n✅ Export concluído: {len(por_centro)} obra(s)")


if __name__ == "__main__":
    main()
