#!/usr/bin/env python3
"""
Migra códigos antigos de centro de custo (001, 054, etc.) para o novo formato YY.NNN
em custos_linhas.xlsx e alocacao_diaria.xlsx.

Mapeamento baseado na lista APROVADOS:
  antigo -> novo (descrição)
"""
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
DADOS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "dados"
CUSTOS_LINHAS = DADOS_PATH / "custos_linhas.xlsx"
ALOCACAO_DIARIA = DADOS_PATH / "alocacao_diaria.xlsx"

# antigo (com ou sem zeros à esquerda) -> novo YY.NNN
MAPEAMENTO = {
    "001": "99.990",   # Sede-Geral -> Manutenções
    "54": "24.54", "054": "24.54",   # Estoril 124
    "80": "24.80", "080": "24.80",   # Moradia Quinta Marinha
    "84": "25.84", "084": "25.84",   # T2 Loja Estoril
    "106": "25.106",                 # SR. LISBOA
    "113": "25.113",                 # CCG
    "115": "25.115",                 # Linda a Velha
    "97": "26.97", "097": "26.97",   # Parede
    "114": "26.114",                 # AV. Elias Garcia 93
    "120": "26.120",                 # Penha de Franca
    "122": "26.122",                 # Barreiro
    "123": "26.123",                 # Tidelli
    "128": "26.128",                 # Nood Colombo
    "129": "26.129",                 # Nood Cascais
    "990": "99.990", "999": "99.990", "0990": "99.990",  # Manutenções
}


def _normalizar(cod: str) -> str:
    return str(cod or "").strip()


def migrar_ficheiro(path: Path, col_centro: str = "centro_custo_codigo") -> tuple[int, list[str]]:
    """Migra centros num ficheiro Excel. Retorna (alterados, nao_mapeados)."""
    if not path.exists():
        return 0, []
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = next((i for i, h in enumerate(headers, 1) if h == col_centro), None)
    if not idx:
        return 0, []
    alterados = 0
    nao_mapeados = []
    for r in range(2, ws.max_row + 1):
        val = _normalizar(ws.cell(r, idx).value)
        if not val or val in MAPEAMENTO.values():
            continue
        # Verificar se já está no formato novo (YY.NNN)
        if "." in val and len(val) >= 5:
            continue
        novo = MAPEAMENTO.get(val) or MAPEAMENTO.get(val.lstrip("0")) or MAPEAMENTO.get(val.zfill(3))
        if novo:
            ws.cell(r, idx).value = novo
            alterados += 1
        else:
            nao_mapeados.append(val)
    wb.save(path)
    return alterados, sorted(set(nao_mapeados))


def main():
    print("Migrar centros de custo antigos -> YY.NNN")
    print()
    tot = 0
    for path, col in [(CUSTOS_LINHAS, "centro_custo_codigo"), (ALOCACAO_DIARIA, "centro_custo_codigo")]:
        if path.exists():
            n, nm = migrar_ficheiro(path, col)
            tot += n
            print(f"  {path.name}: {n} linha(s) atualizada(s)")
            if nm:
                print(f"    Sem mapeamento: {nm}")
        else:
            print(f"  {path.name}: (não existe)")
    print()
    print(f"✅ Total: {tot} centro(s) migrado(s)")


if __name__ == "__main__":
    main()
