#!/usr/bin/env python3
"""
Importa linhas de compras, classifica (materiais/subempreitada) e gera
custos_linhas.xlsx para associação manual de centro de custo.
Regras: fornecedor (principal), IVA (secundário), keywords na descrição (terciário).
"""
import csv
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from utils.taxas_iva import parse_taxa_iva

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
CONFIG_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "config"
DADOS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "dados"
COMPRAS_DOCS = BASE_PATH / "04_COMPRAS" / "compras_documentos.xlsx"
COMPRAS_LINHAS = BASE_PATH / "04_COMPRAS" / "compras_linhas.xlsx"
CENTROS = BASE_PATH / "00_CONFIG" / "centros_custo.xlsx"
CUSTOS_LINHAS = DADOS_PATH / "custos_linhas.xlsx"


def _load_csv_config(name: str, key_col: str, tipo_col: str) -> list[tuple[str, str]]:
    path = CONFIG_PATH / name
    if not path.exists():
        return []
    rows = []
    with open(path, encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            k = (row.get(key_col) or "").strip().lower()
            v = (row.get(tipo_col) or "").strip().lower()
            if k and v and v in ("materiais", "subempreitada", "equipamentos_maquinaria", "custos_sede"):
                rows.append((k, v))
    return rows


def _classificar_linha(
    supplier: str,
    descricao: str,
    tax_pct: float | None,
    fornecedores: list[tuple[str, str]],
    keywords: list[tuple[str, str]],
) -> str:
    sup = (supplier or "").strip().lower()
    desc = (descricao or "").strip().lower()
    for key, tipo in fornecedores:
        if key in sup:
            return tipo
    for key, tipo in keywords:
        if key in desc:
            return tipo
    if tax_pct is not None:
        if abs(tax_pct) < 0.5:
            return "subempreitada"
        if abs(tax_pct - 23) < 1:
            return "materiais"
    return "materiais"


def _load_sheet(wb_path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict]]:
    wb = load_workbook(wb_path, data_only=True)
    ws = wb.active if sheet_name is None else wb[sheet_name]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, i + 1).value for i, h in enumerate(headers) if h}
        rows.append(row)
    return headers, rows


def main() -> None:
    fornecedores = _load_csv_config("classificacao_fornecedores.csv", "fornecedor", "tipo")
    keywords = _load_csv_config("classificacao_keywords.csv", "keyword", "tipo")
    _, docs = _load_sheet(COMPRAS_DOCS)
    _, linhas = _load_sheet(COMPRAS_LINHAS)
    doc_by_id = {str(d.get("id")): d for d in docs if d.get("id")}

    existentes: dict[str, str] = {}
    if CUSTOS_LINHAS.exists():
        _, ex_rows = _load_sheet(CUSTOS_LINHAS)
        for r in ex_rows:
            lid = str(r.get("line_id", ""))
            cc = r.get("centro_custo_codigo")
            if lid and cc:
                existentes[lid] = str(cc).strip()

    out_rows = []
    for ln in linhas:
        doc_id = str(ln.get("rel_document", ""))
        doc = doc_by_id.get(doc_id, {})
        supplier = doc.get("supplier_business_name") or ""
        tax_val = ln.get("tax_percentage")
        tax_pct = parse_taxa_iva(str(tax_val)) if tax_val is not None else None
        tipo = _classificar_linha(
            supplier, ln.get("description") or "", tax_pct, fornecedores, keywords
        )
        line_id = str(ln.get("id", ""))
        centro = existentes.get(line_id, "")
        out_rows.append({
            "line_id": line_id,
            "document_id": doc_id,
            "document_no": doc.get("document_no"),
            "date": doc.get("date"),
            "supplier": supplier,
            "description": ln.get("description"),
            "quantity": ln.get("quantity"),
            "unit_price": ln.get("net_unit_price") or ln.get("unit_price"),
            "net_amount": ln.get("net_amount"),
            "tax_pct": tax_pct,
            "tipo_linha": tipo,
            "centro_custo_codigo": centro,
        })

    DADOS_PATH.mkdir(parents=True, exist_ok=True)
    headers_out = list(out_rows[0].keys()) if out_rows else []
    wb = Workbook()
    ws = wb.active
    ws.title = "custos_linhas"
    for col, h in enumerate(headers_out, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.font = Font(bold=True, color="FFFFFF")
    for r, row in enumerate(out_rows, 2):
        for col, h in enumerate(headers_out, 1):
            ws.cell(row=r, column=col, value=row.get(h))
    wb.save(CUSTOS_LINHAS)
    print(f"✅ custos_linhas: {CUSTOS_LINHAS} ({len(out_rows)} linhas)")
    print("   Preencha 'centro_custo_codigo' e execute exportar_custos_por_obra.py")


if __name__ == "__main__":
    main()
