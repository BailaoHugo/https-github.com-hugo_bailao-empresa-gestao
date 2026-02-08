import os
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))


def read_column_values(xlsx_path: Path, sheet_name: str, header_name: str) -> list[str]:
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    col_idx = headers.get(header_name)
    if not col_idx:
        return []
    values = []
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        value = row[0].value
        if value is not None and str(value).strip():
            values.append(str(value).strip())
    return values


def add_list_validation(ws, ws_lists, col_letter: str, list_row: int, n_vals: int, start_row=2, end_row=2000):
    if n_vals == 0:
        return
    start_cell = ws_lists.cell(row=list_row, column=2).coordinate
    end_cell = ws_lists.cell(row=list_row, column=1 + n_vals).coordinate
    formula = f"=LISTAS!${start_cell[0]}${list_row}:${end_cell[0]}${list_row}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def style_headers(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def build_orcamentos_cabecalho(listas: dict, out_file: Path) -> None:
    wb = Workbook()
    ws_lists = wb.active
    ws_lists.title = "LISTAS"

    row = 1
    for name, values in listas.items():
        ws_lists.cell(row=row, column=1, value=name).font = Font(bold=True)
        for idx, value in enumerate(values, start=2):
            ws_lists.cell(row=row, column=idx, value=value)
        row += 1
    ws_lists.sheet_state = "hidden"

    ws = wb.create_sheet("orcamentos", 0)
    headers = [
        "orcamento_id",
        "centro_custo_codigo",
        "nome_obra",
        "cliente",
        "data_orcamento",
        "versao",
        "estado",
        "moeda",
        "validade_dias",
        "prazo_execucao_dias",
        "margem_prevista_pct",
        "custo_direto_previsto",
        "custo_indireto_previsto",
        "total_previsto",
        "responsavel",
        "origem",
        "notas",
    ]
    style_headers(ws, headers)

    list_row_map = {ws_lists.cell(row=i, column=1).value: i for i in range(1, row)}
    add_list_validation(ws, ws_lists, "B", list_row_map.get("centros_custo", 0), len(listas["centros_custo"]))
    add_list_validation(ws, ws_lists, "D", list_row_map.get("clientes", 0), len(listas["clientes"]))
    add_list_validation(ws, ws_lists, "G", list_row_map.get("estado_orcamento", 0), len(listas["estado_orcamento"]))
    add_list_validation(ws, ws_lists, "H", list_row_map.get("moedas", 0), len(listas["moedas"]))
    add_list_validation(ws, ws_lists, "P", list_row_map.get("origem", 0), len(listas["origem"]))

    for r in range(2, 2001):
        ws[f"E{r}"].number_format = "yyyy-mm-dd"
        ws[f"K{r}"].number_format = "0.00%"
        ws[f"L{r}"].number_format = "#,##0.00"
        ws[f"M{r}"].number_format = "#,##0.00"
        ws[f"N{r}"].number_format = "#,##0.00"

    ws_readme = wb.create_sheet("README")
    today = datetime.now().date().isoformat()
    ws_readme["A1"] = "orcamentos_cabecalho.xlsx — Guia rapido"
    ws_readme["A1"].font = Font(bold=True, size=14)
    ws_readme["A3"] = "1) Cabecalho de orcamentos. Cada orcamento tem um orcamento_id."
    ws_readme["A4"] = "2) O orcamento_id liga as linhas no ficheiro orcamentos_linhas.xlsx."
    ws_readme["A5"] = "3) Nao alterar centro_custo_codigo depois de ter linhas associadas."
    ws_readme["A7"] = f"Gerado em: {today}"
    ws_readme.column_dimensions["A"].width = 110
    for i in range(1, 10):
        ws_readme[f"A{i}"].alignment = Alignment(wrap_text=True)

    wb.save(out_file)


def build_orcamentos_linhas(listas: dict, out_file: Path) -> None:
    wb = Workbook()
    ws_lists = wb.active
    ws_lists.title = "LISTAS"

    row = 1
    for name, values in listas.items():
        ws_lists.cell(row=row, column=1, value=name).font = Font(bold=True)
        for idx, value in enumerate(values, start=2):
            ws_lists.cell(row=row, column=idx, value=value)
        row += 1
    ws_lists.sheet_state = "hidden"

    ws = wb.create_sheet("linhas", 0)
    headers = [
        "orcamento_id",
        "linha_id",
        "rubrica_codigo",
        "descricao",
        "unidade",
        "quantidade",
        "preco_unitario",
        "total_linha",
        "fornecedor_sugerido",
        "item_codigo",
        "origem_preco",
        "data_preco",
        "observacoes",
    ]
    style_headers(ws, headers)

    list_row_map = {ws_lists.cell(row=i, column=1).value: i for i in range(1, row)}
    add_list_validation(ws, ws_lists, "C", list_row_map.get("rubricas", 0), len(listas["rubricas"]))
    add_list_validation(ws, ws_lists, "E", list_row_map.get("unidades", 0), len(listas["unidades"]))
    add_list_validation(ws, ws_lists, "I", list_row_map.get("fornecedores", 0), len(listas["fornecedores"]))
    add_list_validation(ws, ws_lists, "K", list_row_map.get("origem_preco", 0), len(listas["origem_preco"]))

    for r in range(2, 2001):
        ws[f"F{r}"].number_format = "#,##0.00"
        ws[f"G{r}"].number_format = "#,##0.00"
        ws[f"H{r}"].number_format = "#,##0.00"
        ws[f"L{r}"].number_format = "yyyy-mm-dd"

    ws_readme = wb.create_sheet("README")
    today = datetime.now().date().isoformat()
    ws_readme["A1"] = "orcamentos_linhas.xlsx — Guia rapido"
    ws_readme["A1"].font = Font(bold=True, size=14)
    ws_readme["A3"] = "1) Linhas de orcamento. Usa o mesmo orcamento_id do cabecalho."
    ws_readme["A4"] = "2) linha_id pode ser um uuid unico por linha."
    ws_readme["A6"] = f"Gerado em: {today}"
    ws_readme.column_dimensions["A"].width = 110
    for i in range(1, 10):
        ws_readme[f"A{i}"].alignment = Alignment(wrap_text=True)

    wb.save(out_file)


def build_biblioteca_trabalhos(listas: dict, out_file: Path) -> None:
    wb = Workbook()
    ws_lists = wb.active
    ws_lists.title = "LISTAS"

    row = 1
    for name, values in listas.items():
        ws_lists.cell(row=row, column=1, value=name).font = Font(bold=True)
        for idx, value in enumerate(values, start=2):
            ws_lists.cell(row=row, column=idx, value=value)
        row += 1
    ws_lists.sheet_state = "hidden"

    ws = wb.create_sheet("biblioteca", 0)
    headers = [
        "item_id",
        "descricao",
        "rubrica_codigo",
        "unidade",
        "produtividade",
        "preco_unitario_base",
        "fornecedor_preferencial",
        "origem",
        "data_atualizacao",
        "notas",
    ]
    style_headers(ws, headers)

    list_row_map = {ws_lists.cell(row=i, column=1).value: i for i in range(1, row)}
    add_list_validation(ws, ws_lists, "C", list_row_map.get("rubricas", 0), len(listas["rubricas"]))
    add_list_validation(ws, ws_lists, "D", list_row_map.get("unidades", 0), len(listas["unidades"]))
    add_list_validation(ws, ws_lists, "G", list_row_map.get("fornecedores", 0), len(listas["fornecedores"]))
    add_list_validation(ws, ws_lists, "H", list_row_map.get("origem_preco", 0), len(listas["origem_preco"]))

    for r in range(2, 2001):
        ws[f"F{r}"].number_format = "#,##0.00"
        ws[f"I{r}"].number_format = "yyyy-mm-dd"

    ws_readme = wb.create_sheet("README")
    today = datetime.now().date().isoformat()
    ws_readme["A1"] = "biblioteca_trabalhos.xlsx — Guia rapido"
    ws_readme["A1"].font = Font(bold=True, size=14)
    ws_readme["A3"] = "1) Lista de itens reutilizaveis para acelerar orcamentos."
    ws_readme["A4"] = "2) Define rubrica, unidade e preco base."
    ws_readme["A6"] = f"Gerado em: {today}"
    ws_readme.column_dimensions["A"].width = 110
    for i in range(1, 10):
        ws_readme[f"A{i}"].alignment = Alignment(wrap_text=True)

    wb.save(out_file)


def main() -> None:
    centros = read_column_values(
        BASE_PATH / "00_CONFIG" / "centros_custo.xlsx",
        "centros_custo",
        "centro_custo_codigo",
    )
    rubricas = read_column_values(
        BASE_PATH / "00_CONFIG" / "rubricas.xlsx",
        "rubricas",
        "rubrica_codigo",
    )
    moedas = read_column_values(
        BASE_PATH / "00_CONFIG" / "moedas.xlsx",
        "moedas",
        "iso_code",
    )
    unidades = read_column_values(
        BASE_PATH / "00_CONFIG" / "unidades_medida.xlsx",
        "unidades_medida",
        "unit_of_measure",
    )
    clientes = read_column_values(
        BASE_PATH / "01_EMPRESA" / "clientes.xlsx",
        "clientes",
        "business_name",
    )
    fornecedores = read_column_values(
        BASE_PATH / "01_EMPRESA" / "fornecedores.xlsx",
        "fornecedores",
        "business_name",
    )

    listas_cabecalho = {
        "centros_custo": sorted(set(centros)),
        "clientes": sorted(set(clientes)),
        "estado_orcamento": ["Rascunho", "Em Revisao", "Aprovado", "Rejeitado"],
        "moedas": sorted(set(moedas)) or ["EUR"],
        "origem": ["Manual", "Historico", "Template"],
    }

    listas_linhas = {
        "rubricas": sorted(set(rubricas)),
        "unidades": sorted(set(unidades)),
        "fornecedores": sorted(set(fornecedores)),
        "origem_preco": ["Manual", "Historico", "Fornecedor"],
    }

    orc_path = BASE_PATH / "02_OBRAS" / "ORCAMENTOS"
    orc_path.mkdir(parents=True, exist_ok=True)

    build_orcamentos_cabecalho(listas_cabecalho, orc_path / "orcamentos_cabecalho.xlsx")
    build_orcamentos_linhas(listas_linhas, orc_path / "orcamentos_linhas.xlsx")
    build_biblioteca_trabalhos(listas_linhas, orc_path / "biblioteca_trabalhos.xlsx")

    print("✅ Modulo de orcamentos criado.")


if __name__ == "__main__":
    main()
