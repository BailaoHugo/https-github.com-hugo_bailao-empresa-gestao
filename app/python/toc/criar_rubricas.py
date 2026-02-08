import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# Altera apenas se quiseres outra pasta raiz.
BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))

# (Codigo, Nome)
RAW = [
    ("01", "Vendas e Prestacoes Servicos"),
    ("02", "Custos das vendas e Prestacoes Servicos"),
    ("03", "Custos e Fornecimentos Externos"),
    ("04", "Imparidades Provisoes e Depreciacoes"),
    ("05", "Financeiros"),
    ("06", "Impostos"),
    ("07", "Outros"),
    ("08", "Outros Rendimentos"),
]


def add_list_validation(
    ws, ws_lists, col_letter: str, list_row: int, n_vals: int, start_row=2, end_row=2000
) -> None:
    start_cell = ws_lists.cell(row=list_row, column=2).coordinate
    end_cell = ws_lists.cell(row=list_row, column=1 + n_vals).coordinate
    formula = f"=LISTAS!${start_cell[0]}${list_row}:${end_cell[0]}${list_row}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def main() -> None:
    out_file = BASE_PATH / "00_CONFIG" / "rubricas.xlsx"
    out_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws_lists = wb.active
    ws_lists.title = "LISTAS"
    listas = {
        "estado_registo": ["Ativo", "Inativo"],
        "origem": ["TOConline", "Manual"],
        "grupo": ["Centros de custo"],
    }

    row = 1
    for nome, vals in listas.items():
        ws_lists.cell(row=row, column=1, value=nome).font = Font(bold=True)
        for i, v in enumerate(vals, start=2):
            ws_lists.cell(row=row, column=i, value=v)
        row += 1
    ws_lists.sheet_state = "hidden"

    ws = wb.create_sheet("rubricas", 0)

    headers = [
        "rubrica_codigo",
        "rubrica_nome",
        "grupo",
        "estado_registo",
        "origem",
        "data_importacao",
        "notas",
    ]

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    today = datetime.now().date().isoformat()

    for row_idx, (code, name) in enumerate(RAW, start=2):
        values = [
            code,
            name,
            "Centros de custo",
            "Ativo",
            "TOConline",
            today,
            "",
        ]
        for col_idx, v in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    list_row_map = {ws_lists.cell(row=i, column=1).value: i for i in range(1, 1 + len(listas))}
    add_list_validation(ws, ws_lists, "C", list_row_map["grupo"], len(listas["grupo"]))
    add_list_validation(ws, ws_lists, "D", list_row_map["estado_registo"], len(listas["estado_registo"]))
    add_list_validation(ws, ws_lists, "E", list_row_map["origem"], len(listas["origem"]))

    ws_readme = wb.create_sheet("README")
    ws_readme["A1"] = "rubricas.xlsx — Guia rapido"
    ws_readme["A1"].font = Font(bold=True, size=14)
    ws_readme["A3"] = "1) Tabela-mestre de rubricas conforme TOConline."
    ws_readme["A4"] = "2) Nao alterar codigos depois de ligar modulos."
    ws_readme["A6"] = f"Gerado em: {today}"
    ws_readme.column_dimensions["A"].width = 110
    for i in range(1, 10):
        ws_readme[f"A{i}"].alignment = Alignment(wrap_text=True)

    wb.save(out_file)
    print(f"✅ Ficheiro criado: {out_file}")


if __name__ == "__main__":
    main()
