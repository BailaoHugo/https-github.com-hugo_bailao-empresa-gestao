import os
import re
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# Altera apenas se quiseres outra pasta raiz.
BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))

# Lista (Codigo, Descricao) conforme TOConline
# Formato codigo: YY.NNN — 2 digitos = ano inicio obra (25=2025), 3 digitos = numero obra
RAW = [
    ("24.54", "24.54 - ESTORIL 124"),
    ("24.80", "24.80 - MORADIA QUINTA DA MARINHA"),
    ("25.84", "25.84 - T2 LOJA ESTORIL"),
    ("25.106", "25.106 - SR. LISBOA"),
    ("25.113", "25.113 - CCG"),
    ("25.115", "25.115 - LINDA A VELHA"),
    ("26.97", "26.97 - PAREDE"),
    ("26.114", "26.114 - AV. ELIAS GARCIA 93"),
    ("26.120", "26.120 - PENHA DE FRANCA"),
    ("26.122", "26.122 - BARREIRO"),
    ("26.123", "26.123 - TIDELLI"),
    ("26.128", "26.128 - NOOD COLOMBO"),
    ("26.129", "26.129 - NOOD CASCAIS"),
    ("99.990", "99.990 - MANUTENÇOES"),
]


def clean_name(code: str, desc: str) -> str:
    cleaned = re.sub(r"\(\s*inativo\s*\)", "", desc, flags=re.I).strip()
    # Remover codigo do inicio (ex: "25.113 - CCG" ou "001 - Sede")
    cleaned = re.sub(rf"^{re.escape(code)}\s*[-–—]\s*", "", cleaned, flags=re.I).strip()
    if "." not in code and code.isdigit():
        cleaned = re.sub(rf"^\s*0*{int(code)}\s*[-–—]\s*", "", cleaned)
    cleaned = re.sub(r"^\s*\d+\.?\d*\s*[-–—]\s*", "", cleaned).strip()
    return cleaned.strip()


def infer_estado(desc: str) -> str:
    return "Inativo" if "inativo" in desc.lower() else "Ativo"


def infer_tipo(desc: str) -> str:
    dl = desc.lower()
    if "sede" in dl or "geral" in dl:
        return "Sede/Geral"
    if "arrendamento" in dl or "alugado" in dl:
        return "Arrendamento"
    if "pequenas obras" in dl or "reparacoes" in dl:
        return "Pequenas Obras"
    return "Obra"


def infer_subtipo(desc: str) -> str:
    dl = desc.lower()
    if any(k in dl for k in ["loja", "restaurante", "cabeleireiro", "escritorios", "hotel"]):
        return "Comercial"
    if "fabrica" in dl:
        return "Industrial"
    if any(k in dl for k in ["t1", "t2", "t3", "t4", "moradia", "duplex"]):
        return "Habitacional"
    if "sede" in dl or "geral" in dl:
        return "Administrativo"
    if "arrendamento" in dl or "alugado" in dl:
        return "Arrendamento"
    return "Outro"


def add_list_validation(
    ws, ws_lists, col_letter: str, list_row: int, n_vals: int, start_row=2, end_row=2000
) -> None:
    start_cell = ws_lists.cell(row=list_row, column=2).coordinate
    end_cell = ws_lists.cell(row=list_row, column=1 + n_vals).coordinate
    formula = f"=LISTAS!${start_cell[0]}${list_row}:${end_cell[0]}${list_row}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def main() -> None:
    out_file = BASE_PATH / "00_CONFIG" / "centros_custo.xlsx"
    out_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws_lists = wb.active
    ws_lists.title = "LISTAS"
    listas = {
        "estado_registo": ["Ativo", "Inativo"],
        "tipo": ["Obra", "Arrendamento", "Sede/Geral", "Pequenas Obras", "Outro"],
        "subtipo": ["Habitacional", "Comercial", "Administrativo", "Industrial", "Arrendamento", "Outro"],
        "pais": ["Portugal"],
    }

    row = 1
    for nome, vals in listas.items():
        ws_lists.cell(row=row, column=1, value=nome).font = Font(bold=True)
        for i, v in enumerate(vals, start=2):
            ws_lists.cell(row=row, column=i, value=v)
        row += 1
    ws_lists.sheet_state = "hidden"

    ws = wb.create_sheet("centros_custo", 0)

    headers = [
        "cc_uid",
        "centro_custo_codigo",
        "centro_custo_nome",
        "descricao_toconline_original",
        "estado_registo",
        "tipo",
        "subtipo",
        "cliente",
        "local",
        "morada",
        "concelho",
        "distrito",
        "pais",
        "data_inicio",
        "data_fim",
        "gestor_obra",
        "encarregado",
        "contrato_ref",
        "orcamento_ref",
        "valor_contrato_sem_iva",
        "margem_prevista_pct",
        "notas",
        "fonte",
        "data_importacao",
    ]

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    today = datetime.now().date().isoformat()

    namespace = uuid.UUID("5b24e512-2f08-4a0f-9c2b-3cfe28d74a47")

    for row_idx, (code, desc) in enumerate(RAW, start=2):
        uid = str(uuid.uuid5(namespace, code))
        nome = clean_name(code, desc)
        estado = infer_estado(desc)
        tipo = infer_tipo(desc)
        subtipo = infer_subtipo(desc)

        values = [
            uid,
            code,
            nome,
            desc,
            estado,
            tipo,
            subtipo,
            "",
            "",
            "",
            "",
            "",
            "Portugal",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "TOConline",
            today,
        ]
        for col_idx, v in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    list_row_map = {ws_lists.cell(row=i, column=1).value: i for i in range(1, 1 + len(listas))}
    add_list_validation(ws, ws_lists, "E", list_row_map["estado_registo"], len(listas["estado_registo"]))
    add_list_validation(ws, ws_lists, "F", list_row_map["tipo"], len(listas["tipo"]))
    add_list_validation(ws, ws_lists, "G", list_row_map["subtipo"], len(listas["subtipo"]))
    add_list_validation(ws, ws_lists, "M", list_row_map["pais"], len(listas["pais"]))

    for r in range(2, 2001):
        ws[f"N{r}"].number_format = "yyyy-mm-dd"
        ws[f"O{r}"].number_format = "yyyy-mm-dd"
        ws[f"T{r}"].number_format = "#,##0.00"
        ws[f"U{r}"].number_format = "0.00%"

    ws_readme = wb.create_sheet("README")
    ws_readme["A1"] = "centros_custo.xlsx — Guia rapido"
    ws_readme["A1"].font = Font(bold=True, size=14)
    ws_readme["A3"] = "1) Tabela-mestre de Centros de Custo (obras/arrendamentos/geral)."
    ws_readme["A4"] = "2) Nao alterar 'centro_custo_codigo' depois de comecar a ligar modulos."
    ws_readme["A5"] = "3) 'tipo' e 'subtipo' foram pre-preenchidos automaticamente — reve e ajusta."
    ws_readme["A6"] = "4) Campos cliente/local/datas/gestor/contrato servem para cruzar orcamento/planeamento/custos."
    ws_readme["A8"] = f"Gerado em: {today}"
    ws_readme.column_dimensions["A"].width = 110
    for i in range(1, 10):
        ws_readme[f"A{i}"].alignment = Alignment(wrap_text=True)

    wb.save(out_file)
    print(f"✅ Ficheiro criado: {out_file}")


if __name__ == "__main__":
    main()
