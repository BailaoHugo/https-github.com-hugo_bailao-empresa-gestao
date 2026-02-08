import json
import os
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
TOKEN_PATH = Path(os.getenv("TOCONLINE_TOKEN_PATH", "toconline_token.json"))


def load_access_token() -> str:
    if not TOKEN_PATH.exists():
        raise FileNotFoundError(
            f"Nao encontrei {TOKEN_PATH.resolve()}. "
            "Corre primeiro o toconline_oauth.py."
        )
    payload = json.loads(TOKEN_PATH.read_text(encoding="utf-8"))
    token = payload.get("access_token")
    if not token:
        raise RuntimeError("Token nao encontrado em toconline_token.json")
    return token


def get_api_base() -> str:
    return os.getenv("TOCONLINE_API_BASE", "https://api17.toconline.pt").rstrip("/")


def fetch_all(url: str, token: str) -> list[dict]:
    results: list[dict] = []
    api_base = get_api_base()

    while url:
        req = urllib.request.Request(url)
        req.add_header("Authorization", f"Bearer {token}")
        req.add_header("Accept", "application/vnd.api+json")
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        items = data.get("data", [])
        if isinstance(items, dict):
            items = [items]
        results.extend(items)

        next_url = data.get("links", {}).get("next")
        if next_url:
            url = urllib.parse.urljoin(api_base, next_url)
        else:
            url = None

    return results


def normalize_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True)
    return value


def flatten_item(item: dict) -> dict:
    row: dict = {
        "id": item.get("id", ""),
        "type": item.get("type", ""),
    }

    attrs = item.get("attributes", {}) or {}
    for key, value in attrs.items():
        row[key] = normalize_value(value)

    rels = item.get("relationships", {}) or {}
    for rel_name, rel_data in rels.items():
        rel_value = ""
        data = rel_data.get("data") if isinstance(rel_data, dict) else None
        if isinstance(data, list):
            rel_value = ",".join([str(d.get("id")) for d in data if isinstance(d, dict)])
        elif isinstance(data, dict):
            rel_value = data.get("id", "")
        row[f"rel_{rel_name}"] = normalize_value(rel_value)

    return row


def write_excel(rows: list[dict], out_file: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = []
    if rows:
        base_keys = ["id", "type"]
        attr_keys = sorted({k for r in rows for k in r.keys() if k not in base_keys and not k.startswith("rel_")})
        rel_keys = sorted({k for r in rows for k in r.keys() if k.startswith("rel_")})
        headers = base_keys + attr_keys + rel_keys
    else:
        headers = ["id", "type"]

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    today = datetime.now().date().isoformat()

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            value = normalize_value(row.get(key, ""))
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws_readme = wb.create_sheet("README")
    ws_readme["A1"] = f"{out_file.name} — Gerado da API TOConline"
    ws_readme["A1"].font = Font(bold=True, size=14)
    ws_readme["A3"] = "1) Dados exportados diretamente da API."
    ws_readme["A4"] = "2) Nao editar este ficheiro para evitar divergencias."
    ws_readme["A6"] = f"Gerado em: {today}"
    ws_readme.column_dimensions["A"].width = 110
    for i in range(1, 10):
        ws_readme[f"A{i}"].alignment = Alignment(wrap_text=True)

    wb.save(out_file)


def export_resource(token: str, name: str, path: str, out_file: Path) -> None:
    api_base = get_api_base()
    url = f"{api_base}{path}"
    items = fetch_all(url, token)
    rows = [flatten_item(item) for item in items]
    write_excel(rows, out_file, name)
    print(f"✅ {name}: {out_file}")


def main() -> None:
    token = load_access_token()

    resources = [
        # 00_CONFIG
        ("rubricas", "/api/expense_categories", BASE_PATH / "00_CONFIG" / "rubricas.xlsx"),
        ("unidades_medida", "/api/units_of_measure", BASE_PATH / "00_CONFIG" / "unidades_medida.xlsx"),
        ("familias_itens", "/api/item_families", BASE_PATH / "00_CONFIG" / "familias_itens.xlsx"),
        ("taxas", "/api/taxes", BASE_PATH / "00_CONFIG" / "taxas.xlsx"),
        ("paises", "/api/countries", BASE_PATH / "00_CONFIG" / "paises.xlsx"),
        ("moedas", "/api/currencies", BASE_PATH / "00_CONFIG" / "moedas.xlsx"),
        ("series_documentos", "/api/commercial_document_series", BASE_PATH / "00_CONFIG" / "series_documentos.xlsx"),
        ("descritores_taxa", "/api/tax_descriptors", BASE_PATH / "00_CONFIG" / "descritores_taxa.xlsx"),
        # 01_EMPRESA
        ("clientes", "/api/customers", BASE_PATH / "01_EMPRESA" / "clientes.xlsx"),
        ("fornecedores", "/api/suppliers", BASE_PATH / "01_EMPRESA" / "fornecedores.xlsx"),
        ("contactos", "/api/contacts", BASE_PATH / "01_EMPRESA" / "contactos.xlsx"),
        ("moradas", "/api/addresses", BASE_PATH / "01_EMPRESA" / "moradas.xlsx"),
        ("contas_bancarias", "/api/bank_accounts", BASE_PATH / "01_EMPRESA" / "contas_bancarias.xlsx"),
        ("contas_caixa", "/api/cash_accounts", BASE_PATH / "01_EMPRESA" / "contas_caixa.xlsx"),
        # 04_COMPRAS
        ("compras_documentos", "/api/commercial_purchases_documents", BASE_PATH / "04_COMPRAS" / "compras_documentos.xlsx"),
        ("compras_linhas", "/api/commercial_purchases_document_lines", BASE_PATH / "04_COMPRAS" / "compras_linhas.xlsx"),
        ("compras_pagamentos", "/api/commercial_purchases_payments", BASE_PATH / "04_COMPRAS" / "compras_pagamentos.xlsx"),
        ("compras_pagamentos_linhas", "/api/commercial_purchases_payment_lines", BASE_PATH / "04_COMPRAS" / "compras_pagamentos_linhas.xlsx"),
        # 05_VENDAS
        ("vendas_documentos", "/api/commercial_sales_documents", BASE_PATH / "05_VENDAS" / "vendas_documentos.xlsx"),
        ("vendas_linhas", "/api/commercial_sales_document_lines", BASE_PATH / "05_VENDAS" / "vendas_linhas.xlsx"),
        ("vendas_recibos", "/api/commercial_sales_receipts", BASE_PATH / "05_VENDAS" / "vendas_recibos.xlsx"),
        ("vendas_recibos_linhas", "/api/commercial_sales_receipt_lines", BASE_PATH / "05_VENDAS" / "vendas_recibos_linhas.xlsx"),
    ]

    for name, path, out_file in resources:
        out_file.parent.mkdir(parents=True, exist_ok=True)
        export_resource(token, name, path, out_file)


if __name__ == "__main__":
    main()
