import json
import os
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
TOKEN_PATH = Path(os.getenv("TOCONLINE_TOKEN_PATH", "toconline_token.json"))


def load_access_token() -> str:
    if not TOKEN_PATH.exists():
        raise FileNotFoundError(
            f"Nao encontrei {TOKEN_PATH.resolve()}. "
            "Corre primeiro o toconline_oauth.py."
        )
    payload = json.loads(TOKEN_PATH.read_text(encoding="utf-8"))
    token = payload.get("access_token")
    if not token:
        raise RuntimeError("Token nao encontrado em toconline_token.json")
    return token


def get_api_base() -> str:
    return os.getenv("TOCONLINE_API_BASE", "https://api17.toconline.pt").rstrip("/")


def fetch_all_expense_categories(token: str) -> list[dict]:
    api_base = get_api_base()
    url = f"{api_base}/api/expense_categories"
    results: list[dict] = []

    while url:
        req = urllib.request.Request(url)
        req.add_header("Authorization", f"Bearer {token}")
        req.add_header("Accept", "application/vnd.api+json")
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        page_items = data.get("data", [])
        if isinstance(page_items, dict):
            page_items = [page_items]
        results.extend(page_items)

        next_url = data.get("links", {}).get("next")
        if next_url:
            url = urllib.parse.urljoin(api_base, next_url)
        else:
            url = None

    return results


def build_excel(rows: list[dict], out_file: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "rubricas"

    headers = [
        "rubrica_id",
        "rubrica_codigo",
        "rubrica_nome",
        "parent_rubrica_id",
        "tax_code",
        "tax_deductibility",
        "accounting_number",
        "accounting_suffix",
        "is_main",
        "is_deletable",
        "is_editable",
        "is_ready",
        "categories_count",
        "help",
        "grouping_id",
        "profit_and_losses_id",
        "created_at",
        "updated_at",
        "fonte",
        "data_importacao",
    ]

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    today = datetime.now().date().isoformat()

    for row_idx, item in enumerate(rows, start=2):
        attrs = item.get("attributes", {})
        values = [
            item.get("id", ""),
            attrs.get("accounting_number") or attrs.get("accounting_suffix") or "",
            attrs.get("name", ""),
            attrs.get("parent_expense_category_id", ""),
            attrs.get("tax_code", ""),
            attrs.get("tax_deductibility", ""),
            attrs.get("accounting_number", ""),
            attrs.get("accounting_suffix", ""),
            attrs.get("is_main", ""),
            attrs.get("is_deletable", ""),
            attrs.get("is_editable", ""),
            attrs.get("is_ready", ""),
            attrs.get("categories_count", ""),
            attrs.get("help", ""),
            attrs.get("enumerations_expense_category_grouping_id", ""),
            attrs.get("enumerations_profit_and_losses_category_id", ""),
            attrs.get("created_at", ""),
            attrs.get("updated_at", ""),
            "TOConline",
            today,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(out_file)


def main() -> None:
    token = load_access_token()
    rows = fetch_all_expense_categories(token)

    out_file = BASE_PATH / "00_CONFIG" / "rubricas.xlsx"
    out_file.parent.mkdir(parents=True, exist_ok=True)

    build_excel(rows, out_file)
    print(f"✅ Ficheiro criado: {out_file}")


if __name__ == "__main__":
    main()
