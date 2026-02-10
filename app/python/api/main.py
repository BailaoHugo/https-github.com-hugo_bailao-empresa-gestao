#!/usr/bin/env python3
"""
API para registo de despesas por foto.
POST /api/registar-despesa: foto + centro_custo_codigo → OCR → gravar custos
"""
import os
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from utils.taxas_iva import parse_taxa_iva

BASE_PATH = Path(os.getenv("GESTAO_BASE_PATH", "/home/bailan/empresa-gestao/GESTAO_EMPRESA"))
DADOS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "dados"
CUSTOS_LINHAS = DADOS_PATH / "custos_linhas.xlsx"
CUSTOS_REGISTO = DADOS_PATH / "custos_registo.xlsx"
OBRAS_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA" / "obras"
UPLOADS_PATH = DADOS_PATH / "uploads"
FACTURAS_EXTRAIDAS = DADOS_PATH / "facturas_extraidas"
UPLOADS_PATH.mkdir(parents=True, exist_ok=True)
FACTURAS_EXTRAIDAS.mkdir(parents=True, exist_ok=True)

try:
    from fastapi import FastAPI, File, Form, UploadFile, HTTPException, Query
    from fastapi.middleware.cors import CORSMiddleware
except ImportError:
    print("Instale: pip install fastapi uvicorn python-multipart")
    sys.exit(1)

try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    XL_AVAILABLE = True
except ImportError:
    XL_AVAILABLE = False

app = FastAPI(title="Registo Despesas")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def _ocr_image(img_path: Path) -> str:
    if not OCR_AVAILABLE:
        return ""
    try:
        img = Image.open(img_path)
        if img.mode != "RGB":
            img = img.convert("RGB")
        return pytesseract.image_to_string(img, lang="por+eng")
    except Exception as e:
        return f"[OCR erro: {e}]"


def _extrair_dados_ocr(texto: str) -> dict:
    """Extrai fornecedor, data, valores e IVA do texto OCR."""
    out = {"supplier": "", "date": "", "net_amount": None, "tax_pct": None, "description": ""}
    lines = [l.strip() for l in texto.splitlines() if l.strip()]
    for i, line in enumerate(lines):
        # Data (dd-mm-yyyy, dd/mm/yyyy, yyyy-mm-dd)
        m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", line)
        if m and not out["date"]:
            d, mo, y = m.groups()
            out["date"] = f"20{y[-2:]}-{mo.zfill(2)}-{d.zfill(2)}" if len(y) == 2 else f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
        # Valores: Total, IVA
        for p in [r"total[:\s]*(\d+[,.]\d{2})", r"(\d+[,.]\d{2})\s*€", r"iva[:\s]*(\d+[,.]?\d*)%?"]:
            m = re.search(p, line.lower())
            if m:
                v = m.group(1).replace(",", ".")
                try:
                    n = float(v)
                    if "iva" in line.lower() or "%" in line:
                        out["tax_pct"] = parse_taxa_iva(v)
                    elif out["net_amount"] is None and n > 0:
                        out["net_amount"] = n
                except ValueError:
                    pass
    # Primeira linha costuma ser fornecedor
    if lines and not out["supplier"]:
        out["supplier"] = lines[0][:100]
    return out


def _append_custo(centro: str, dados: dict, origem: str = "foto") -> None:
    if not XL_AVAILABLE:
        return
    if not CUSTOS_LINHAS.exists():
        DADOS_PATH.mkdir(parents=True, exist_ok=True)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "custos_linhas"
        headers = ["line_id", "document_id", "document_no", "date", "supplier", "description",
                   "quantity", "unit_price", "net_amount", "tax_pct", "tipo_linha", "centro_custo_codigo"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.font = Font(bold=True, color="FFFFFF")
        wb.save(CUSTOS_LINHAS)
    wb = load_workbook(CUSTOS_LINHAS)
    ws = wb.active
    next_row = ws.max_row + 1
    row_data = {
        "line_id": f"foto_{next_row}",
        "document_id": "",
        "document_no": origem,
        "date": dados.get("date"),
        "supplier": dados.get("supplier"),
        "description": dados.get("description") or f"Registo por {origem}",
        "quantity": 1,
        "unit_price": dados.get("net_amount"),
        "net_amount": dados.get("net_amount"),
        "tax_pct": dados.get("tax_pct"),
        "tipo_linha": "materiais",
        "centro_custo_codigo": centro,
    }
    headers = [c.value for c in ws[1]]
    for col, h in enumerate(headers, 1):
        if h in row_data:
            ws.cell(row=next_row, column=col, value=row_data[h])
    wb.save(CUSTOS_LINHAS)


@app.get("/api/centros-custo")
def listar_centros():
    """Lista centros de custo para o dropdown."""
    path = BASE_PATH / "00_CONFIG" / "centros_custo.xlsx"
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    idx_cod = next((i for i, c in enumerate(ws[1], 1) if c.value == "centro_custo_codigo"), None)
    idx_nome = next((i for i, c in enumerate(ws[1], 1) if c.value == "centro_custo_nome"), None)
    if not idx_cod:
        return []
    out = []
    for r in range(2, ws.max_row + 1):
        cod = ws.cell(r, idx_cod).value
        nome = ws.cell(r, idx_nome).value if idx_nome else ""
        if cod:
            out.append({"codigo": str(cod), "nome": str(nome or cod)})
    return out


ORCAMENTOS_CABECALHO = BASE_PATH / "02_OBRAS" / "ORCAMENTOS" / "orcamentos_cabecalho.xlsx"


@app.get("/api/orcamentos")
def listar_orcamentos():
    """Lista orçamentos do ficheiro Excel."""
    if not XL_AVAILABLE or not ORCAMENTOS_CABECALHO.exists():
        return []
    try:
        wb = load_workbook(ORCAMENTOS_CABECALHO, data_only=True)
        ws = None
        for name in ["orcamentos", "orcamentos_cabecalho", "cabecalho"]:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        idx_id = next((i for i, h in enumerate(headers, 1) if "orcamento_id" in h.lower() or h == "orcamento_id"), None)
        idx_obra = next((i for i, h in enumerate(headers, 1) if "nome_obra" in h.lower() or "obra" in h.lower()), None)
        idx_cliente = next((i for i, h in enumerate(headers, 1) if "cliente" in h.lower()), None)
        idx_data = next((i for i, h in enumerate(headers, 1) if "data" in h.lower()), None)
        idx_estado = next((i for i, h in enumerate(headers, 1) if "estado" in h.lower()), None)
        idx_total = next((i for i, h in enumerate(headers, 1) if "total" in h.lower()), None)
        if not idx_id:
            return []
        out = []
        for r in range(2, ws.max_row + 1):
            oid = ws.cell(r, idx_id).value
            if not oid:
                continue
            out.append({
                "orcamento_id": str(oid),
                "nome_obra": str(ws.cell(r, idx_obra).value or "") if idx_obra else "",
                "cliente": str(ws.cell(r, idx_cliente).value or "") if idx_cliente else "",
                "data_orcamento": str(ws.cell(r, idx_data).value or "") if idx_data else "",
                "estado": str(ws.cell(r, idx_estado).value or "") if idx_estado else "",
                "total_previsto": ws.cell(r, idx_total).value if idx_total else None,
            })
        return out
    except Exception:
        return []


@app.post("/api/registar-despesa")
async def registar_despesa(
    centro_custo_codigo: str = Form(...),
    file: UploadFile = File(...),
):
    """
    Recebe foto de fatura/recibo + centro de custo.
    Faz OCR, extrai dados e grava em custos_linhas.
    """
    centro = centro_custo_codigo.strip()
    if not centro:
        raise HTTPException(400, "centro_custo_codigo obrigatório")

    ext = Path(file.filename or "").suffix.lower() or ".jpg"
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
        raise HTTPException(400, "Formato inválido. Use jpg, png ou webp.")

    content = await file.read()
    save_path = UPLOADS_PATH / f"{os.urandom(8).hex()}{ext}"
    save_path.write_bytes(content)

    texto = _ocr_image(save_path)
    dados = _extrair_dados_ocr(texto)
    dados["description"] = dados.get("description") or file.filename or "Foto"

    # Extrair factura estruturada e guardar em facturas_extraidas (igual ao fluxo email)
    ficheiro_extraido = None
    try:
        from custos.extrair_factura import (
            extrair_factura,
            guardar_factura_json,
            guardar_factura_excel,
        )
        origem = file.filename or save_path.name
        factura = extrair_factura(texto, origem=f"foto:{origem}|centro:{centro}")
        base_name = save_path.stem + "_extraida"
        guardar_factura_json(factura, FACTURAS_EXTRAIDAS / f"{base_name}.json")
        guardar_factura_excel(factura, FACTURAS_EXTRAIDAS / f"{base_name}.xlsx")
        ficheiro_extraido = f"{base_name}.xlsx"
    except Exception:
        pass  # continua e grava custos_linhas mesmo que extrair_factura falhe

    _append_custo(centro, dados, origem=file.filename or "foto")

    return {
        "ok": True,
        "centro_custo_codigo": centro,
        "ocr_texto": texto[:500] if texto else "(OCR não disponível)",
        "dados_extraidos": dados,
        "ficheiro_extraido": ficheiro_extraido,
    }


def _load_custos_registo() -> list[dict]:
    """Carrega custos_registo.xlsx."""
    if not XL_AVAILABLE or not CUSTOS_REGISTO.exists():
        return []
    wb = load_workbook(CUSTOS_REGISTO, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    return [
        {h: ws.cell(r, i + 1).value for i, h in enumerate(headers) if h}
        for r in range(2, ws.max_row + 1)
    ]


@app.get("/api/custos/obras")
def listar_obras_com_custos():
    """Lista obras (centros de custo) com totais de custos por tipo."""
    rows = _load_custos_registo()
    por_obra: dict[str, dict] = {}
    tipos = ("subempreitadas", "materiais", "mao_obra", "equipamentos_maquinaria", "custos_sede")
    for r in rows:
        cc = str(r.get("centro_custo_codigo") or "").strip()
        if not cc:
            continue
        if cc not in por_obra:
            por_obra[cc] = {"centro_custo_codigo": cc, "total": 0, **{t: 0 for t in tipos}}
        t = (r.get("tipo_linha") or "materiais").strip().lower()
        if t == "subempreitada":
            t = "subempreitadas"
        if t not in tipos:
            t = "materiais"
        val = r.get("net_amount") or r.get("unit_price") or 0
        try:
            val = float(val) if val is not None else 0
        except (TypeError, ValueError):
            val = 0
        por_obra[cc][t] = por_obra[cc][t] + val
        por_obra[cc]["total"] = por_obra[cc]["total"] + val
    return list(por_obra.values())


@app.get("/api/custos/obras/{centro}")
def custos_por_obra(centro: str, tipo: str | None = None, capitulo: str | None = None):
    """Lista custos de uma obra, opcionalmente filtrados por tipo e capítulo."""
    rows = _load_custos_registo()
    filtrado = [r for r in rows if str(r.get("centro_custo_codigo") or "").strip() == centro]
    if tipo:
        t = tipo.strip().lower()
        if t == "subempreitada":
            t = "subempreitadas"
        filtrado = [r for r in filtrado if (r.get("tipo_linha") or "").strip().lower() in (t, "subempreitada" if t == "subempreitadas" else t)]
    if capitulo:
        filtrado = [r for r in filtrado if (r.get("capitulo_orcamento") or "").strip() == capitulo]
    return {"centro_custo_codigo": centro, "linhas": filtrado, "total_linhas": len(filtrado)}


@app.get("/api/custos/capitulos")
def listar_capitulos_orcamento():
    """Lista capítulos do orçamento para filtro/dropdown."""
    path = BASE_PATH / "00_CONFIG" / "capitulos_orcamento.xlsx"
    if not XL_AVAILABLE or not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx_id = next((i for i, h in enumerate(headers, 1) if "id" in str(h).lower()), None)
    idx_nome = next((i for i, h in enumerate(headers, 1) if "nome" in str(h).lower()), None)
    if not idx_id:
        return []
    out = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, idx_id).value
        nome = ws.cell(r, idx_nome).value if idx_nome else ""
        if cid:
            out.append({"id": str(cid), "nome": str(nome or cid)})
    return out


# --- Base de Dados endpoints ---

CONFIG_PATH = BASE_PATH / "00_CONFIG"
EMPRESA_PATH = BASE_PATH / "01_EMPRESA"
CONTAB_PATH = BASE_PATH / "03_CONTABILIDADE_ANALITICA"
CLASSIFICACAO_PATH = CONTAB_PATH / "config" / "classificacao_fornecedores.csv"


def _load_excel_as_dicts(path: Path) -> list[dict]:
    """Carrega ficheiro Excel e devolve lista de dicionários (1ª linha = headers)."""
    if not XL_AVAILABLE or not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [h for h in (c.value for c in ws[1]) if h]
    return [
        {h: ws.cell(r, i + 1).value for i, h in enumerate(headers) if i < len(headers)}
        for r in range(2, ws.max_row + 1)
    ]


def _load_classificacao_fornecedores() -> dict[str, str]:
    """Mapa fornecedor -> tipo (materiais/subempreitada)."""
    out: dict[str, str] = {}
    if not CLASSIFICACAO_PATH.exists():
        return out
    with open(CLASSIFICACAO_PATH, encoding="utf-8") as f:
        lines = [l.strip() for l in f if l.strip()]
    for line in lines[1:]:  # skip header
        parts = [p.strip() for p in line.split(",")]
        if len(parts) >= 2:
            out[parts[0].strip()] = parts[1].strip()
    return out


def _apply_text_filter(rows: list[dict], q, text_cols: list[str]) -> list[dict]:
    """Filtra linhas onde alguma coluna de texto contenha q (case-insensitive)."""
    ql = (q.strip().lower()) if isinstance(q, str) and q.strip() else ""
    if not ql:
        return rows
    return [r for r in rows if any(
        ql in str(r.get(c) or "").lower() for c in text_cols if r.get(c)
    )]


def _opt_str(v):
    """Normaliza parâmetro opcional de query para string ou None."""
    return v.strip() if isinstance(v, str) and v else None


def _opt_float(v):
    """Normaliza parâmetro opcional de query para float ou None."""
    if v is None or not isinstance(v, (int, float, str)):
        return None
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _to_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _apply_value_filters(rows: list[dict], valor_min, valor_max, col: str = "net_amount") -> list[dict]:
    """Filtra por valor_min e valor_max na coluna indicada."""
    out = rows
    vmin = _opt_float(valor_min)
    vmax = _opt_float(valor_max)
    if vmin is not None:
        out = [r for r in out if _to_float(r.get(col)) >= vmin]
    if vmax is not None:
        out = [r for r in out if _to_float(r.get(col)) <= vmax]
    return out


def _apply_date_filters(rows: list[dict], data_inicio, data_fim, col: str = "date") -> list[dict]:
    """Filtra por intervalo de datas (strings yyyy-mm-dd)."""
    out = rows
    di = _opt_str(data_inicio)
    df = _opt_str(data_fim)
    if di:
        out = [r for r in out if str(r.get(col) or "") >= di]
    if df:
        out = [r for r in out if str(r.get(col) or "") <= df]
    return out


@app.get("/api/base-dados/entidades")
def listar_entidades():
    """Lista metadata das entidades disponíveis na base de dados."""
    return [
        {"id": "centros-custo", "nome": "Centros de custo", "filtros": ["q"]},
        {"id": "fornecedores", "nome": "Fornecedores", "filtros": ["q"]},
        {"id": "clientes", "nome": "Clientes / Empresas", "filtros": ["q"]},
        {"id": "materiais", "nome": "Materiais", "filtros": ["q", "centro", "data_inicio", "data_fim", "valor_min", "valor_max"]},
        {"id": "subempreiteiros", "nome": "Subempreiteiros", "filtros": ["q", "centro", "data_inicio", "data_fim", "valor_min", "valor_max"]},
        {"id": "custos", "nome": "Custos (registo completo)", "filtros": ["q", "tipo", "centro", "data_inicio", "data_fim", "valor_min", "valor_max"]},
        {"id": "trabalhadores", "nome": "Trabalhadores", "filtros": ["q"]},
        {"id": "capitulos", "nome": "Capítulos orçamento", "filtros": ["q"]},
    ]


@app.get("/api/base-dados/centros-custo")
def base_dados_centros_custo(q: str | None = Query(None, description="Pesquisa textual")):
    rows = listar_centros()
    data = [{"centro_custo_codigo": r["codigo"], "centro_custo_nome": r["nome"]} for r in rows]
    data = _apply_text_filter(data, q, ["centro_custo_codigo", "centro_custo_nome"])
    return {"dados": data, "total": len(data)}


@app.get("/api/base-dados/fornecedores")
def base_dados_fornecedores(q: str | None = Query(None, description="Pesquisa textual")):
    path = EMPRESA_PATH / "fornecedores.xlsx"
    rows = _load_excel_as_dicts(path)
    rows = _apply_text_filter(rows, q, ["business_name", "id", "internal_observations"])
    return {"dados": rows, "total": len(rows)}


@app.get("/api/base-dados/clientes")
def base_dados_clientes(q: str | None = Query(None, description="Pesquisa textual")):
    path = EMPRESA_PATH / "clientes.xlsx"
    rows = _load_excel_as_dicts(path)
    rows = _apply_text_filter(rows, q, ["business_name", "id", "contact_name"])
    return {"dados": rows, "total": len(rows)}


@app.get("/api/base-dados/materiais")
def base_dados_materiais(
    q: str | None = Query(None),
    centro: str | None = Query(None),
    data_inicio: str | None = Query(None),
    data_fim: str | None = Query(None),
    valor_min: float | None = Query(None),
    valor_max: float | None = Query(None),
):
    rows = _load_custos_registo()
    rows = [r for r in rows if (str(r.get("tipo_linha") or "").strip().lower() in ("materiais", "material"))]
    centro_s = _opt_str(centro)
    if centro_s:
        rows = [r for r in rows if str(r.get("centro_custo_codigo") or "").strip() == centro_s]
    rows = _apply_date_filters(rows, data_inicio, data_fim)
    rows = _apply_value_filters(rows, valor_min, valor_max)
    rows = _apply_text_filter(rows, q, ["supplier", "description", "document_no"])
    return {"dados": rows, "total": len(rows), "soma_net_amount": sum(_to_float(r.get("net_amount")) for r in rows)}


@app.get("/api/base-dados/subempreiteiros")
def base_dados_subempreiteiros(
    q: str | None = Query(None),
    centro: str | None = Query(None),
    data_inicio: str | None = Query(None),
    data_fim: str | None = Query(None),
    valor_min: float | None = Query(None),
    valor_max: float | None = Query(None),
):
    rows = _load_custos_registo()
    rows = [r for r in rows if (str(r.get("tipo_linha") or "").strip().lower() in ("subempreitadas", "subempreitada", "subempreiteiros"))]
    clf = _load_classificacao_fornecedores()
    for r in rows:
        r["tipo_classificado"] = clf.get(str(r.get("supplier") or ""), "")
    centro_s = _opt_str(centro)
    if centro_s:
        rows = [r for r in rows if str(r.get("centro_custo_codigo") or "").strip() == centro_s]
    rows = _apply_date_filters(rows, data_inicio, data_fim)
    rows = _apply_value_filters(rows, valor_min, valor_max)
    rows = _apply_text_filter(rows, q, ["supplier", "description", "document_no"])
    return {"dados": rows, "total": len(rows), "soma_net_amount": sum(_to_float(r.get("net_amount")) for r in rows)}


@app.get("/api/base-dados/custos")
def base_dados_custos(
    q: str | None = Query(None),
    tipo: str | None = Query(None),
    centro: str | None = Query(None),
    data_inicio: str | None = Query(None),
    data_fim: str | None = Query(None),
    valor_min: float | None = Query(None),
    valor_max: float | None = Query(None),
):
    rows = _load_custos_registo()
    tipo_s = _opt_str(tipo)
    if tipo_s:
        t = tipo_s.lower()
        if t == "subempreitada":
            t = "subempreitadas"
        rows = [r for r in rows if (str(r.get("tipo_linha") or "").strip().lower() == t)]
    centro_s = _opt_str(centro)
    if centro_s:
        rows = [r for r in rows if str(r.get("centro_custo_codigo") or "").strip() == centro_s]
    rows = _apply_date_filters(rows, data_inicio, data_fim)
    rows = _apply_value_filters(rows, valor_min, valor_max)
    rows = _apply_text_filter(rows, q, ["supplier", "description", "document_no", "centro_custo_codigo"])
    soma = sum(_to_float(r.get("net_amount")) or _to_float(r.get("unit_price")) for r in rows)
    return {"dados": rows, "total": len(rows), "soma_net_amount": soma}


@app.get("/api/base-dados/trabalhadores")
def base_dados_trabalhadores(q: str | None = Query(None)):
    path = DADOS_PATH / "trabalhadores.xlsx"
    rows = _load_excel_as_dicts(path)
    rows = _apply_text_filter(rows, q, ["codigo", "nome", "origem"])
    return {"dados": rows, "total": len(rows)}


@app.get("/api/base-dados/capitulos")
def base_dados_capitulos(q: str | None = Query(None)):
    rows = listar_capitulos_orcamento()
    data = [{"capitulo_id": r["id"], "capitulo_nome": r["nome"]} for r in rows]
    data = _apply_text_filter(data, q, ["capitulo_id", "capitulo_nome"])
    return {"dados": data, "total": len(data)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
